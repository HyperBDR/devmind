from __future__ import annotations

import base64
import binascii
import fcntl
import os
import signal
import subprocess
import textwrap
from collections.abc import Iterator
from contextlib import contextmanager
from copy import copy
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from hashlib import sha256
from io import BytesIO
from math import ceil
from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile, ZipInfo

from django.conf import settings
from django.db import IntegrityError, transaction
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as SpreadsheetImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import (
    absolute_coordinate,
    coordinate_to_tuple,
    get_column_letter,
    quote_sheetname,
)
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.writer.excel import ExcelWriter
from PIL import Image as PillowImage
from PIL import UnidentifiedImageError
from quotation.models import QuotationTemplate, QuotationTemplateStatus
from quotation.services.storage import (
    delete_document,
    resolve_document_path,
    template_storage_key,
    write_document_atomic,
)

LEGACY_DEFAULT_TEMPLATE_NAME = "DevMind standard quotation"
DEFAULT_TEMPLATE_NAME = "DevMind managed standard quotation"
DEFAULT_TEMPLATE_VERSION = 2
CURRENT_RENDERER_VERSION = "quotation-preview-xlsx-v7"
DEFAULT_WORKSHEET = "Quotation"


def estimate_wrapped_lines(value: str, *, width: int) -> int:
    """Estimate rendered lines for wrapped text in a spreadsheet cell."""
    safe_width = max(1, width)
    lines = value.splitlines() or [""]
    return sum(
        max(1, len(textwrap.wrap(line, width=safe_width)) if line else 1)
        for line in lines
    )
REQUIRED_TEMPLATE_NAMES = {
    "billing_company",
    "billing_contact",
    "billing_email",
    "client_company",
    "contact_person",
    "currency",
    "email",
    "expire_date",
    "grand_total",
    "issuer_company_name",
    "issuer_contact_email",
    "issuer_contact_name",
    "issuer_signature",
    "line_items_start",
    "payment_terms",
    "project_name",
    "quote_date",
    "quote_no",
    "remarks_disclaimer",
    "subtotal_before_vat",
    "vat_amount",
}
OPTIONAL_TEMPLATE_NAMES = {"tax_label", "vat_rate"}


class TemplateValidationError(ValueError):
    def __init__(self, message: str, *, code: str):
        super().__init__(message)
        self.code = code


class PdfConversionError(RuntimeError):
    def __init__(self, message: str, *, code: str, retryable: bool = True):
        super().__init__(message)
        self.code = code
        self.retryable = retryable


class PdfConversionBusyError(PdfConversionError):
    pass


class PdfConversionTimeoutError(PdfConversionError, TimeoutError):
    pass


def _description_row_height(description: str) -> float:
    """Return enough height for wrapped description text."""
    lines = str(description or "").splitlines() or [""]
    line_count = sum(max(1, ceil(len(line) / 24)) for line in lines)
    return min(120, max(24, 6 + line_count * 15))


@contextmanager
def _libreoffice_conversion_slot() -> Iterator[None]:
    """Acquire one cross-process LibreOffice conversion slot."""
    lock_dir = Path(settings.QUOTATION_RENDER_LOCK_DIR)
    lock_dir.mkdir(parents=True, exist_ok=True)
    for slot in range(settings.QUOTATION_RENDER_CONCURRENCY):
        lock_path = lock_dir / f"slot-{slot}.lock"
        lock_file = lock_path.open("a+b")
        try:
            fcntl.flock(
                lock_file.fileno(),
                fcntl.LOCK_EX | fcntl.LOCK_NB,
            )
        except BlockingIOError:
            lock_file.close()
            continue
        try:
            yield
        finally:
            fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)
            lock_file.close()
        return
    raise PdfConversionBusyError(
        "LibreOffice conversion capacity is busy",
        code="libreoffice_busy",
    )


def _save_workbook_deterministic(workbook: Workbook) -> bytes:
    """Serialize an XLSX without wall-clock metadata or ZIP timestamps."""
    stable_modified = (
        workbook.properties.modified
        or workbook.properties.created
        or datetime(2000, 1, 1)
    )
    workbook.properties.modified = stable_modified.replace(microsecond=0)
    raw_output = BytesIO()
    archive = ZipFile(
        raw_output,
        "w",
        ZIP_DEFLATED,
        allowZip64=True,
    )
    ExcelWriter(workbook, archive).save()

    canonical_output = BytesIO()
    with ZipFile(BytesIO(raw_output.getvalue())) as source:
        with ZipFile(
            canonical_output,
            "w",
            ZIP_DEFLATED,
            allowZip64=True,
        ) as target:
            for source_info in source.infolist():
                target_info = ZipInfo(
                    source_info.filename,
                    date_time=(1980, 1, 1, 0, 0, 0),
                )
                target_info.compress_type = source_info.compress_type
                target_info.create_system = source_info.create_system
                target_info.external_attr = source_info.external_attr
                target_info.internal_attr = source_info.internal_attr
                target_info.comment = source_info.comment
                target.writestr(
                    target_info,
                    source.read(source_info.filename),
                )
    return canonical_output.getvalue()


def _add_defined_name(
    workbook: Workbook,
    name: str,
    cell_reference: str,
) -> None:
    workbook.defined_names.add(
        DefinedName(
            name,
            attr_text=f"'{DEFAULT_WORKSHEET}'!${cell_reference}",
        )
    )


def _build_managed_template_bytes(*, version: int) -> bytes:
    """Build one managed template version for creation or identification."""
    if version not in {1, DEFAULT_TEMPLATE_VERSION}:
        raise ValueError("unsupported managed template version")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = DEFAULT_WORKSHEET
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 9
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 15
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 16
    sheet.column_dimensions["G"].width = 17

    thin = Side(style="thin", color="94A3B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="E2E8F0")

    sheet.merge_cells("A1:G1")
    sheet["A1"] = "OnePro Cloud Limited"
    sheet["A1"].font = Font(size=18, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:G2")
    sheet["A2"] = "Quotation"
    sheet["A2"].font = Font(size=16, bold=True, underline="single")
    sheet["A2"].alignment = Alignment(horizontal="center")

    labels = {
        "A4": "Quote No.",
        "F4": "Date",
        "F5": "Valid Till",
        "A6": "Ship to",
        "C6": "Contact",
        "E6": "Email",
        "A7": "Bill to",
        "C7": "Contact",
        "E7": "Email",
        "A8": "Project",
        "C8": "Payment Terms",
        "F8": "Currency",
    }
    for coordinate, label in labels.items():
        sheet[coordinate] = label
        sheet[coordinate].font = Font(bold=True)

    headers = [
        "Line",
        "Description",
        "Qty",
        "List Price",
        "Discount",
        "Net Unit Price",
        "Extended Price",
    ]
    for column, label in enumerate(headers, 1):
        cell = sheet.cell(row=10, column=column, value=label)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        item_cell = sheet.cell(row=11, column=column)
        item_cell.border = border
        item_cell.alignment = Alignment(
            horizontal="left" if column == 2 else "right",
            vertical="top",
            wrap_text=True,
        )

    sheet["F12"] = "Subtotal"
    if version == 1:
        sheet["F13"] = "Tax"
    else:
        sheet["E13"] = "Tax"
        sheet["F13"] = "Rate"
    sheet["F14"] = "Grand Total"
    for row in range(12, 15):
        if version >= DEFAULT_TEMPLATE_VERSION:
            sheet.cell(row=row, column=5).font = Font(bold=True)
            sheet.cell(row=row, column=5).border = border
        sheet.cell(row=row, column=6).font = Font(bold=True)
        sheet.cell(row=row, column=7).font = Font(bold=True)
        sheet.cell(row=row, column=6).border = border
        sheet.cell(row=row, column=7).border = border

    sheet["A16"] = "Remarks"
    sheet["A16"].font = Font(bold=True)
    sheet.merge_cells("B16:G16")
    sheet["B16"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet["A18"] = "Prepared by"
    sheet["A18"].font = Font(bold=True)
    sheet["D18"] = "Email"
    sheet["D18"].font = Font(bold=True)
    sheet["D20"] = "Signature"
    sheet["D20"].font = Font(bold=True)
    sheet.merge_cells("E20:G22")

    names = {
        "issuer_company_name": "A1",
        "quote_no": "B4",
        "quote_date": "G4",
        "expire_date": "G5",
        "client_company": "B6",
        "contact_person": "D6",
        "email": "G6",
        "billing_company": "B7",
        "billing_contact": "D7",
        "billing_email": "G7",
        "project_name": "B8",
        "payment_terms": "D8",
        "currency": "G8",
        "line_items_start": "A11",
        "subtotal_before_vat": "G12",
        "vat_amount": "G13",
        "grand_total": "G14",
        "remarks_disclaimer": "B16",
        "issuer_contact_name": "B18",
        "issuer_contact_email": "E18",
        "issuer_signature": "E20",
    }
    if version >= DEFAULT_TEMPLATE_VERSION:
        names.update(
            {
                "tax_label": "E13",
                "vat_rate": "F13",
            }
        )
    for name, coordinate in names.items():
        _add_defined_name(workbook, name, coordinate)

    sheet.freeze_panes = "A10"
    sheet.print_area = "A1:G22"
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    stable_timestamp = datetime(2000, 1, 1)
    workbook.properties.created = stable_timestamp
    workbook.properties.modified = stable_timestamp
    return _save_workbook_deterministic(workbook)


def build_default_template_bytes() -> bytes:
    """Build the managed fallback template used on fresh deployments."""
    return _build_managed_template_bytes(version=DEFAULT_TEMPLATE_VERSION)


def _template_fingerprint(content: bytes) -> str:
    """Hash workbook semantics while ignoring generated timestamps."""
    workbook = load_workbook(BytesIO(content), read_only=False)
    try:
        stable_timestamp = datetime(2000, 1, 1)
        workbook.properties.created = stable_timestamp
        workbook.properties.modified = stable_timestamp
        normalized = _save_workbook_deterministic(workbook)
    finally:
        workbook.close()
    return sha256(normalized).hexdigest()


def _is_legacy_managed_template(template: QuotationTemplate) -> bool:
    """Identify the shipped v1 template by immutable workbook content."""
    if template.version != 1 or template.name not in {
        LEGACY_DEFAULT_TEMPLATE_NAME,
        DEFAULT_TEMPLATE_NAME,
    }:
        return False
    try:
        content = template_path(template).read_bytes()
        expected = _build_managed_template_bytes(version=1)
        return _template_fingerprint(content) == _template_fingerprint(expected)
    except (OSError, TemplateValidationError):
        return False


def validate_template_bytes(content: bytes) -> None:
    max_bytes = settings.QUOTATION_MAX_TEMPLATE_BYTES
    if not content.startswith(b"PK\x03\x04"):
        raise TemplateValidationError(
            "Quotation template is not an XLSX file",
            code="template_invalid_signature",
        )
    if len(content) > max_bytes:
        raise TemplateValidationError(
            "Quotation template exceeds the size limit",
            code="template_too_large",
        )
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            member_names = {member.filename for member in members}
    except BadZipFile as exc:
        raise TemplateValidationError(
            "Quotation template is not a valid ZIP container",
            code="template_invalid_zip",
        ) from exc
    expanded_bytes = sum(member.file_size for member in members)
    if expanded_bytes > settings.QUOTATION_MAX_TEMPLATE_EXPANDED_BYTES:
        raise TemplateValidationError(
            "Quotation template expands beyond the size limit",
            code="template_expanded_too_large",
        )
    if "xl/vbaProject.bin" in member_names:
        raise TemplateValidationError(
            "Macro-enabled quotation templates are not allowed",
            code="template_macros_forbidden",
        )
    if any(name.startswith("xl/externalLinks/") for name in member_names):
        raise TemplateValidationError(
            "External workbook links are not allowed",
            code="template_external_links_forbidden",
        )
    if "xl/connections.xml" in member_names:
        raise TemplateValidationError(
            "External workbook connections are not allowed",
            code="template_external_connections_forbidden",
        )
    try:
        workbook = load_workbook(BytesIO(content), read_only=False)
    except Exception as exc:
        raise TemplateValidationError(
            "Quotation template cannot be opened",
            code="template_unreadable",
        ) from exc
    has_worksheet = DEFAULT_WORKSHEET in workbook.sheetnames
    available_names = set(workbook.defined_names)
    workbook.close()
    if not has_worksheet:
        raise TemplateValidationError(
            "Quotation template is missing the Quotation worksheet",
            code="template_worksheet_missing",
        )
    missing = sorted(REQUIRED_TEMPLATE_NAMES - available_names)
    if missing:
        message = "Quotation template is missing named ranges: "
        message += ", ".join(missing)
        raise TemplateValidationError(
            message,
            code="template_named_ranges_missing",
        )


def ensure_default_template(*, created_by=None) -> QuotationTemplate:
    active = (
        QuotationTemplate.objects.filter(status=QuotationTemplateStatus.ACTIVE)
        .order_by("-version", "id")
        .first()
    )
    if active is not None and not _is_legacy_managed_template(active):
        return active

    content = build_default_template_bytes()
    validate_template_bytes(content)
    template = QuotationTemplate(
        name=DEFAULT_TEMPLATE_NAME,
        version=DEFAULT_TEMPLATE_VERSION,
        content_hash=sha256(content).hexdigest(),
        status=QuotationTemplateStatus.ACTIVE,
        created_by=created_by,
    )
    template.storage_key = template_storage_key(template.id)
    write_document_atomic(content, template.storage_key)
    try:
        with transaction.atomic():
            current = (
                QuotationTemplate.objects.select_for_update()
                .filter(status=QuotationTemplateStatus.ACTIVE)
                .order_by("-version", "id")
                .first()
            )
            if current is not None and not _is_legacy_managed_template(current):
                selected = current
            else:
                existing = (
                    QuotationTemplate.objects.select_for_update()
                    .filter(
                        name=DEFAULT_TEMPLATE_NAME,
                        version=DEFAULT_TEMPLATE_VERSION,
                    )
                    .first()
                )
                if (
                    existing is not None
                    and existing.content_hash != template.content_hash
                ):
                    raise TemplateValidationError(
                        "Managed default template version conflicts with "
                        "different content",
                        code="default_template_version_conflict",
                    )
                if current is not None:
                    current.status = QuotationTemplateStatus.ARCHIVED
                    current.save(
                        update_fields=["status", "updated_at"],
                    )
                if existing is not None:
                    existing.status = QuotationTemplateStatus.ACTIVE
                    existing.save(
                        update_fields=["status", "updated_at"],
                    )
                    selected = existing
                else:
                    template.save(force_insert=True)
                    selected = template
    except IntegrityError:
        delete_document(template.storage_key)
        active = (
            QuotationTemplate.objects.filter(
                status=QuotationTemplateStatus.ACTIVE,
            )
            .order_by("-version", "id")
            .first()
        )
        if active is not None:
            return active
        raise
    except Exception:
        delete_document(template.storage_key)
        raise
    if selected.pk != template.pk:
        delete_document(template.storage_key)
    return selected


def register_template_version(
    *,
    name: str,
    version: int,
    content: bytes,
    status: str,
    created_by=None,
) -> QuotationTemplate:
    """Validate and atomically register an immutable XLSX template."""
    validate_template_bytes(content)
    template = QuotationTemplate(
        name=name,
        version=version,
        content_hash=sha256(content).hexdigest(),
        status=status,
        created_by=created_by,
    )
    template.storage_key = template_storage_key(template.id)
    write_document_atomic(content, template.storage_key)
    try:
        with transaction.atomic():
            if status == QuotationTemplateStatus.ACTIVE:
                QuotationTemplate.objects.filter(
                    status=QuotationTemplateStatus.ACTIVE,
                ).update(status=QuotationTemplateStatus.ARCHIVED)
            template.save(force_insert=True)
    except Exception:
        delete_document(template.storage_key)
        raise
    return template


def template_path(template: QuotationTemplate) -> Path:
    path = resolve_document_path(template.storage_key)
    if not path.is_file():
        raise TemplateValidationError(
            "Quotation template file is missing",
            code="template_file_missing",
        )
    content = path.read_bytes()
    validate_template_bytes(content)
    if sha256(content).hexdigest() != template.content_hash:
        raise TemplateValidationError(
            "Quotation template hash does not match its version",
            code="template_hash_mismatch",
        )
    return path


def _defined_cell(workbook, name: str):
    definition = workbook.defined_names.get(name)
    destinations = list(definition.destinations) if definition else []
    if len(destinations) != 1:
        raise TemplateValidationError(
            f"Named range {name} must point to one cell",
            code="template_named_range_invalid",
        )
    sheet_name, coordinate = destinations[0]
    return workbook[sheet_name], coordinate.replace("$", "")


def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(row=source_row, column=column)
        target = sheet.cell(row=target_row, column=column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
    source_height = sheet.row_dimensions[source_row].height
    sheet.row_dimensions[target_row].height = source_height


def _insert_rows_preserving_layout(
    workbook,
    sheet,
    row: int,
    amount: int,
) -> None:
    """Insert template rows and move layout metadata openpyxl leaves stale."""
    shifted_dimensions = []
    for dimension_row, dimension in list(sheet.row_dimensions.items()):
        if dimension_row < row:
            continue
        shifted = copy(dimension)
        shifted.index = dimension_row + amount
        shifted_dimensions.append((dimension_row, shifted))

    shifted_merges = []
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.max_row < row:
            continue
        current = CellRange(str(merged_range))
        sheet.unmerge_cells(str(merged_range))
        if current.min_row >= row:
            current.shift(row_shift=amount)
        else:
            current.max_row += amount
        shifted_merges.append(str(current))

    sheet.insert_rows(row, amount=amount)
    for dimension_row, _dimension in shifted_dimensions:
        del sheet.row_dimensions[dimension_row]
    for _dimension_row, dimension in shifted_dimensions:
        sheet.row_dimensions[dimension.index] = dimension
    for merged_range in shifted_merges:
        sheet.merge_cells(merged_range)

    for drawing in [*sheet._images, *sheet._charts]:
        anchor = drawing.anchor
        if isinstance(anchor, str):
            anchor_row, anchor_column = coordinate_to_tuple(anchor)
            if anchor_row >= row:
                drawing.anchor = (
                    f"{get_column_letter(anchor_column)}" f"{anchor_row + amount}"
                )
            continue
        for marker_name in ("_from", "to"):
            marker = getattr(anchor, marker_name, None)
            if marker is not None and marker.row >= row - 1:
                marker.row += amount

    template_names = REQUIRED_TEMPLATE_NAMES | OPTIONAL_TEMPLATE_NAMES
    for name in template_names:
        definition = workbook.defined_names.get(name)
        destinations = list(definition.destinations) if definition else []
        if len(destinations) != 1:
            continue
        sheet_name, coordinate = destinations[0]
        if sheet_name != sheet.title:
            continue
        target = CellRange(coordinate.replace("$", ""))
        if target.min_row < row:
            continue
        target.shift(row_shift=amount)
        sheet_reference = quote_sheetname(sheet.title)
        cell_reference = absolute_coordinate(str(target))
        definition.attr_text = f"{sheet_reference}!{cell_reference}"


def _signature_image(data_url: str) -> SpreadsheetImage | None:
    if not data_url:
        return None
    try:
        header, encoded = data_url.split(",", 1)
    except ValueError as exc:
        raise TemplateValidationError(
            "Quotation signature is not a data URL",
            code="signature_invalid",
        ) from exc
    allowed_headers = {
        "data:image/jpeg;base64",
        "data:image/jpg;base64",
        "data:image/png;base64",
    }
    if header.lower() not in allowed_headers:
        raise TemplateValidationError(
            "Quotation signature must be a PNG or JPEG data URL",
            code="signature_type_invalid",
        )
    max_encoded_bytes = ((settings.QUOTATION_MAX_SIGNATURE_BYTES + 2) // 3) * 4
    if len(encoded) > max_encoded_bytes:
        raise TemplateValidationError(
            "Quotation signature exceeds the size limit",
            code="signature_too_large",
        )
    try:
        image_bytes = base64.b64decode(encoded, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise TemplateValidationError(
            "Quotation signature contains invalid base64 data",
            code="signature_invalid",
        ) from exc
    if len(image_bytes) > settings.QUOTATION_MAX_SIGNATURE_BYTES:
        raise TemplateValidationError(
            "Quotation signature exceeds the size limit",
            code="signature_too_large",
        )
    try:
        with PillowImage.open(BytesIO(image_bytes)) as image:
            image.verify()
        with PillowImage.open(BytesIO(image_bytes)) as image:
            width, height = image.size
            image_format = image.format
    except (OSError, UnidentifiedImageError) as exc:
        raise TemplateValidationError(
            "Quotation signature image is invalid",
            code="signature_invalid",
        ) from exc
    if image_format not in {"JPEG", "PNG"} or width > 4000 or height > 4000:
        raise TemplateValidationError(
            "Quotation signature image dimensions are invalid",
            code="signature_dimensions_invalid",
        )
    image = SpreadsheetImage(BytesIO(image_bytes))
    scale = min(180 / image.width, 60 / image.height, 1)
    image.width = round(image.width * scale)
    image.height = round(image.height * scale)
    return image


def _image_anchor(
    *,
    row: int,
    column: int,
    width: int,
    height: int,
) -> TwoCellAnchor:
    """Return an image anchor supported by Excel-compatible web viewers."""
    emu_per_pixel = 9525
    return TwoCellAnchor(
        editAs="oneCell",
        _from=AnchorMarker(row=row, col=column),
        to=AnchorMarker(
            row=row,
            col=column,
            rowOff=height * emu_per_pixel,
            colOff=width * emu_per_pixel,
        ),
    )


def render_quotation_xlsx(
    template: QuotationTemplate,
    snapshot: dict,
) -> bytes:
    """Render the spreadsheet using the live quotation preview layout."""
    del template
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = DEFAULT_WORKSHEET
    sheet.sheet_view.showGridLines = False
    widths = (12, 24, 8, 12, 10, 17, 17)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    thin = Side(style="thin", color="CBD5E1")
    dark = Side(style="thin", color="0F172A")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    dark_border = Border(left=dark, right=dark, top=dark, bottom=dark)
    bottom_border = Border(bottom=dark)
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    muted_fill = PatternFill("solid", fgColor="F8FAFC")
    no_fill = PatternFill(fill_type=None)
    font = Font(name="Arial", size=10, color="0F172A")
    bold = Font(name="Arial", size=10, bold=True, color="0F172A")

    def style_range(row: int, start: int = 1, end: int = 7, **kwargs):
        for column in range(start, end + 1):
            cell = sheet.cell(row=row, column=column)
            cell.font = kwargs.get("font", font)
            cell.fill = kwargs.get("fill", no_fill)
            cell.border = kwargs.get("border", Border())
            cell.alignment = kwargs.get(
                "alignment",
                Alignment(vertical="center", wrap_text=True),
            )

    def merged(row: int, start: int, end: int, value="", **kwargs):
        sheet.merge_cells(
            start_row=row,
            start_column=start,
            end_row=row,
            end_column=end,
        )
        cell = sheet.cell(row=row, column=start, value=value)
        style_range(row, start, end, **kwargs)
        cell.font = kwargs.get("font", font)
        cell.alignment = kwargs.get(
            "alignment", Alignment(vertical="center", wrap_text=True)
        )
        return cell

    def value(key: str, fallback=""):
        result = snapshot.get(key)
        return fallback if result in (None, "") else result

    is_imported_quotation = snapshot.get("source_type") == "document_import"

    def issuer_value(key: str, fallback=""):
        if is_imported_quotation:
            return value(key)
        return value(key, fallback)

    def number_text(value_):
        if value_ in (None, ""):
            return ""
        return format(Decimal(str(value_)).normalize(), "f")

    def numeric_value(value_):
        if value_ in (None, ""):
            return None
        rounded = Decimal(str(value_)).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
        if rounded == rounded.to_integral():
            return int(rounded)
        return float(rounded)

    def numeric_format(value_, *, grouped=False, suffix=""):
        rounded = Decimal(str(value_ or 0)).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
        places = max(-rounded.normalize().as_tuple().exponent, 0)
        pattern = "#,##0" if grouped else "0"
        if places:
            pattern += "." + ("0" * places)
        if suffix:
            pattern += f'"{suffix}"'
        return pattern

    def money(value_):
        if value_ in (None, "", 0, 0.0):
            return None
        amount = Decimal(str(value_))
        if not amount:
            return None
        return numeric_value(amount)

    for row in range(1, 100):
        sheet.row_dimensions[row].height = 18

    logo_path = Path(__file__).resolve().parent.parent / "assets" / (
        "onepro-logo.png"
    )
    logo = SpreadsheetImage(logo_path)
    logo_ratio = logo.height / logo.width
    logo.width = 132
    logo.height = round(132 * logo_ratio)
    logo.anchor = _image_anchor(
        row=0,
        column=0,
        width=logo.width,
        height=logo.height,
    )
    sheet.add_image(logo)

    sheet.row_dimensions[1].height = 60
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].height = 28
    sheet.row_dimensions[4].height = 15
    merged(1, 1, 7, "")
    merged(
        2,
        1,
        7,
        value("issuer_company_name"),
        font=Font(name="Arial", size=18, bold=True, color="0F172A"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    merged(
        3,
        1,
        7,
        "Quotation",
        font=Font(
            name="Arial",
            size=22,
            bold=True,
            underline="single",
            color="0F172A",
        ),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    merged(4, 1, 7, "")
    merged(5, 1, 7, "")

    right_details = (
        (6, "Date:", value("quote_date")),
        (7, "Quote No.:", value("quote_no")),
        (8, "Quote Valid Till:", value("expire_date")),
    )
    for row, label, content in right_details:
        sheet.cell(row, 6, label)
        sheet.cell(row, 7, content)
        style_range(
            row,
            6,
            6,
            font=bold,
            alignment=Alignment(horizontal="right", vertical="center"),
        )
        style_range(
            row,
            7,
            7,
            border=bottom_border,
            alignment=Alignment(horizontal="right", vertical="center"),
        )

    merged(
        7,
        1,
        2,
        "Ship to",
        font=bold,
        fill=header_fill,
        border=dark_border,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    customer_details = (
        (8, "Company :", value("client_company")),
        (9, "Name :", value("contact_person")),
        (10, "Email :", value("email")),
    )
    for row, label, content in customer_details:
        merged(
            row,
            1,
            2,
            f"{label} {content}",
            border=dark_border,
        )
    merged(11, 1, 7, "")
    sheet.row_dimensions[11].height = 9
    merged(
        12,
        1,
        2,
        "Bill to:",
        font=bold,
        fill=header_fill,
        border=dark_border,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    billing_details = (
        (
            13,
            "Company :",
            issuer_value("billing_company", value("client_company")),
        ),
        (
            14,
            "Name :",
            issuer_value("billing_contact", value("contact_person")),
        ),
        (15, "Email :", issuer_value("billing_email", value("email"))),
    )
    for row, label, content in billing_details:
        merged(
            row,
            1,
            2,
            f"{label} {content}",
            border=dark_border,
        )
    merged(16, 1, 7, "")
    sheet.row_dimensions[16].height = 12
    merged(17, 1, 7, "")
    sheet.row_dimensions[17].height = 12

    meta_headers = (
        "Contact Person",
        "Email",
        "Project",
        "Payment Terms",
        "Currency",
    )
    meta_values = [
        issuer_value("issuer_contact_name", value("contact_person")),
        issuer_value("issuer_contact_email", value("email")),
        issuer_value("project_name", "-"),
        issuer_value("payment_terms", "-"),
        value("currency"),
    ]
    meta_positions = ((1, 1), (2, 2), (3, 5), (6, 6), (7, 7))
    meta_font = Font(name="Arial", size=9, bold=True, color="0F172A")
    for content, (start, end) in zip(meta_headers, meta_positions):
        merged(
            18,
            start,
            end,
            content,
            font=meta_font,
            fill=header_fill,
            border=cell_border,
        )
    sheet.cell(19, 1, meta_values[0])
    sheet.cell(19, 2, meta_values[1])
    merged(19, 3, 5, meta_values[2])
    sheet.cell(19, 6, meta_values[3])
    sheet.cell(19, 7, meta_values[4])
    style_range(19, border=cell_border)
    for cell in sheet[19]:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    merged(20, 1, 7, "")
    sheet.row_dimensions[20].height = 15

    items = list(snapshot.get("items") or [])
    groups = (
        (
            "Software",
            [item for item in items if item.get("type") == "Software"],
            1,
            snapshot.get("software_subtotal"),
        ),
        (
            "Others",
            [item for item in items if item.get("type") != "Software"],
            1,
            snapshot.get("others_subtotal"),
        ),
    )
    row = 21
    headers = (
        "Item",
        "Description",
        "Qty",
        "List Price",
        "Discount (%)",
        "Discounted Price",
        "Extended Price",
    )
    for section, section_items, minimum, subtotal in groups:
        merged(
            row,
            1,
            7,
            section,
            font=Font(name="Arial", size=11, bold=True),
            fill=header_fill,
            border=cell_border,
        )
        row += 1
        for column, header in enumerate(headers, 1):
            sheet.cell(row, column, header)
        style_range(
            row,
            font=bold,
            fill=muted_fill,
            border=cell_border,
            alignment=Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            ),
        )
        row += 1
        rows = section_items + [
            {} for _ in range(max(minimum - len(section_items), 0))
        ]
        for index, item in enumerate(rows, 1):
            description = item.get("description") or item.get("name") or ""
            values = [
                index if description else "",
                description,
                numeric_value(item.get("qty")) if description else None,
                money(item.get("list_price")) if description else None,
                (
                    numeric_value(item.get("discount_percent") or 0)
                    if description
                    else None
                ),
                money(item.get("net_unit_price")) if description else None,
                money(item.get("extended_price")) if description else None,
            ]
            for column, content in enumerate(values, 1):
                sheet.cell(row, column, content)
            style_range(row, border=cell_border)
            sheet.row_dimensions[row].height = _description_row_height(
                description,
            )
            for column in (1, 3, 5):
                sheet.cell(row, column).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            sheet.cell(row, 3).number_format = numeric_format(
                item.get("qty"),
            )
            sheet.cell(row, 5).number_format = numeric_format(
                item.get("discount_percent"),
                suffix="%",
            )
            for column in (4, 6, 7):
                sheet.cell(row, column).alignment = Alignment(
                    horizontal="right",
                    vertical="center",
                )
            for column, field in (
                (4, "list_price"),
                (6, "net_unit_price"),
                (7, "extended_price"),
            ):
                sheet.cell(row, column).number_format = numeric_format(
                    item.get(field),
                    grouped=True,
                )
            row += 1
        for column in range(1, 5):
            sheet.cell(row, column).border = Border()
        sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        subtotal_label = (
            "Software subscription subtotal:"
            if section == "Software"
            else "Others Subtotal:"
        )
        sheet.cell(row, 5, subtotal_label)
        sheet.cell(row, 7, money(subtotal))
        sheet.cell(row, 7).number_format = numeric_format(
            subtotal,
            grouped=True,
        )
        style_range(
            row,
            5,
            7,
            font=bold,
            border=cell_border,
            alignment=Alignment(horizontal="right", vertical="center"),
        )
        row += 1
        merged(row, 1, 7, "")
        sheet.row_dimensions[row].height = 15
        row += 1

    totals = (
        (
            f"Subtotal before {value('tax_label')}:",
            snapshot.get("subtotal_before_vat"),
        ),
        (
            f"{value('tax_label')} Amount "
            f"({number_text(value('vat_rate', 0))}%):",
            snapshot.get("vat_amount"),
        ),
        ("Grand Total:", snapshot.get("grand_total")),
    )
    for label, amount in totals:
        sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        sheet.cell(row, 5, label)
        sheet.cell(row, 7, money(amount))
        sheet.cell(row, 7).number_format = numeric_format(
            amount,
            grouped=True,
        )
        style_range(
            row,
            5,
            7,
            font=bold,
            border=cell_border,
            alignment=Alignment(horizontal="right", vertical="center"),
        )
        row += 1
    merged(row, 1, 7, "")
    sheet.row_dimensions[row].height = 9
    row += 1
    merged(row, 1, 7, "")
    sheet.row_dimensions[row].height = 9
    row += 1
    merged(row, 1, 7, "Additional Notes & Disclaimers:", font=bold)
    row += 1
    merged(
        row,
        1,
        7,
        value("remarks_disclaimer"),
        font=Font(name="Arial", size=9, color="334155"),
        border=cell_border,
        fill=muted_fill,
        alignment=Alignment(vertical="top", wrap_text=True),
    )
    notes_lines = estimate_wrapped_lines(
        str(value("remarks_disclaimer")),
        width=100,
    )
    sheet.row_dimensions[row].height = max(30, notes_lines * 12)
    row += 1
    for _ in range(2):
        merged(row, 1, 7, "")
        sheet.row_dimensions[row].height = 12
        row += 1
    merged(
        row,
        1,
        7,
        "To indicate Customer acceptance of this quotation, please sign "
        "below and return one copy of this quotation to OnePro Cloud.",
    )
    sheet.row_dimensions[row].height = 24
    row += 1
    merged(row, 1, 7, "")
    sheet.row_dimensions[row].height = 24
    row += 1
    merged(row, 1, 3, "")
    merged(row, 4, 4, "")
    merged(row, 5, 7, value("issuer_company_name"), font=bold)
    row += 1
    signature_row = row
    merged(
        row,
        1,
        3,
        "________________________",
        alignment=Alignment(vertical="bottom"),
    )
    merged(
        row,
        5,
        7,
        "________________________",
        alignment=Alignment(vertical="bottom"),
    )
    sheet.row_dimensions[row].height = 30
    signature = _signature_image(str(value("issuer_signature")))
    if signature is not None:
        signature.anchor = _image_anchor(
            row=signature_row - 1,
            column=4,
            width=signature.width,
            height=signature.height,
        )
        sheet.add_image(signature)
    row += 1
    merged(row, 1, 3, "Name :")
    merged(
        row,
        5,
        7,
        "Name : "
        f"{issuer_value('issuer_contact_name', value('contact_person'))}",
    )
    row += 1
    merged(row, 1, 3, "Title :")
    merged(
        row,
        5,
        7,
        f"Title : {issuer_value('issuer_contact_title', 'Sales Manager')}",
    )
    row += 1
    merged(row, 1, 3, "Email :")
    merged(
        row,
        5,
        7,
        f"Email : {issuer_value('issuer_contact_email', value('email'))}",
    )
    sheet.print_area = f"A1:G{row}"
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    workbook.properties.created = datetime(2000, 1, 1)
    workbook.properties.modified = datetime(2000, 1, 1)
    content = _save_workbook_deterministic(workbook)
    workbook.close()
    generated = load_workbook(BytesIO(content), read_only=True)
    try:
        if DEFAULT_WORKSHEET not in generated.sheetnames:
            raise TemplateValidationError(
                "Generated quotation worksheet is missing",
                code="generated_worksheet_missing",
            )
    finally:
        generated.close()
    return content


def _terminate_process_group(process: subprocess.Popen) -> None:
    try:
        os.killpg(process.pid, signal.SIGTERM)
    except ProcessLookupError:
        return
    try:
        process.wait(timeout=3)
    except subprocess.TimeoutExpired:
        try:
            os.killpg(process.pid, signal.SIGKILL)
        except ProcessLookupError:
            pass
        process.wait(timeout=3)


def convert_xlsx_to_pdf(excel_bytes: bytes, *, job_id: str) -> bytes:
    """Convert XLSX bytes while respecting shared process capacity."""
    with _libreoffice_conversion_slot():
        return _convert_xlsx_to_pdf_unlocked(excel_bytes, job_id=job_id)


def _convert_xlsx_to_pdf_unlocked(excel_bytes: bytes, *, job_id: str) -> bytes:
    """Convert XLSX bytes with an isolated headless LibreOffice profile."""
    with TemporaryDirectory(prefix=f"quotation-render-{job_id}-") as root:
        root_path = Path(root)
        profile_path = root_path / "profile"
        output_path = root_path / "output"
        profile_path.mkdir()
        output_path.mkdir()
        input_path = root_path / "quotation.xlsx"
        input_path.write_bytes(excel_bytes)
        command = [
            settings.QUOTATION_SOFFICE_BINARY,
            "--headless",
            "--nologo",
            "--nodefault",
            "--nofirststartwizard",
            f"-env:UserInstallation={profile_path.as_uri()}",
            "--convert-to",
            "pdf",
            "--outdir",
            str(output_path),
            str(input_path),
        ]
        try:
            process = subprocess.Popen(
                command,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                start_new_session=True,
            )
        except FileNotFoundError as exc:
            raise PdfConversionError(
                "LibreOffice executable is unavailable",
                code="libreoffice_unavailable",
            ) from exc
        try:
            _stdout, stderr = process.communicate(
                timeout=settings.QUOTATION_RENDER_TIMEOUT_SECONDS
            )
        except subprocess.TimeoutExpired as exc:
            _terminate_process_group(process)
            raise PdfConversionTimeoutError(
                "LibreOffice conversion timed out",
                code="libreoffice_timeout",
            ) from exc
        if process.returncode != 0:
            error_type = "conversion_failed" if stderr else "process_failed"
            raise PdfConversionError(
                "LibreOffice conversion failed",
                code=f"libreoffice_{error_type}",
            )
        pdf_path = output_path / "quotation.pdf"
        if not pdf_path.is_file():
            raise PdfConversionError(
                "LibreOffice produced no PDF file",
                code="libreoffice_no_output",
            )
        pdf_bytes = pdf_path.read_bytes()
        if not pdf_bytes.startswith(b"%PDF-"):
            raise PdfConversionError(
                "LibreOffice produced an invalid PDF file",
                code="libreoffice_invalid_pdf",
                retryable=False,
            )
        if len(pdf_bytes) > settings.QUOTATION_MAX_PDF_BYTES:
            raise PdfConversionError(
                "Generated PDF exceeds the size limit",
                code="libreoffice_pdf_too_large",
                retryable=False,
            )
        return pdf_bytes
