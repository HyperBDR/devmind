from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pypdf import PdfReader

from quotation.models import DocumentAsset
from quotation.services.document_parsing.excel_parser import _decimal
from quotation.services.document_parsing.schemas import (
    ParsedDocumentData,
    ParsedQuotation,
    ParsedQuotationItem,
)


def _document_date(asset: DocumentAsset) -> date:
    stem = Path(asset.file_name).stem
    patterns = (
        (r"(?<!\d)(20\d{6})(?!\d)", "%Y%m%d"),
        (r"(?<!\d)(\d{2}[.-]\d{2}[.-]\d{2})(?!\d)", "%d.%m.%y"),
    )
    for pattern, date_format in patterns:
        match = re.search(pattern, stem)
        if not match:
            continue
        raw = match.group(1).replace("-", ".")
        try:
            return datetime.strptime(raw, date_format).date()
        except ValueError:
            continue
    return asset.created_at.date() if asset.created_at else date.today()


def _safe_source_number(asset: DocumentAsset) -> str:
    stem = Path(asset.file_name).stem
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", stem).strip("_.-")
    return (cleaned or f"DOC_{asset.id[:8]}")[:120]


def _number(value: str) -> Decimal:
    return _decimal(re.sub(r"(?<=\d)\s+(?=\d)", "", value))


def _pdf_layout(path: Path) -> str:
    try:
        import pymupdf

        with pymupdf.open(path) as document:
            text = "\n".join(
                page.get_text("text", sort=True) for page in document
            )
        if text.strip():
            return text
    except (ImportError, RuntimeError, ValueError):
        pass
    reader = PdfReader(str(path))
    return "\n".join(
        page.extract_text(extraction_mode="layout") or ""
        for page in reader.pages
    )


def _pdf_label(layout: str, pattern: str) -> str:
    match = re.search(pattern, layout, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", match.group(1)).strip() if match else ""


def _pdf_date(layout: str, label: str) -> date | None:
    raw = _pdf_label(
        layout,
        rf"{re.escape(label)}\s*:?\s*([0-9A-Za-z.-]+)",
    )
    for date_format in (
        "%d.%m.%y",
        "%d-%b-%y",
        "%d-%b-%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(raw, date_format).date()
        except ValueError:
            continue
    return None


def _pdf_items(layout: str) -> list[ParsedQuotationItem]:
    from quotation.services.document_parsing.pdf_parser import (
        _parse_currency_item_line,
    )

    items = []
    pending_description = ""
    for line in layout.splitlines():
        stripped = line.strip()
        if len(stripped) > 1000:
            continue
        if stripped.count("$") >= 3 and re.match(r"^\d+\s", stripped):
            item = _parse_currency_item_line(
                re.sub(r"\s+", " ", stripped),
                pending_description,
            )
            if item is not None:
                item.line_no = len(items) + 1
                item.type = "Software" if not items else "Other"
                items.append(item)
                pending_description = ""
                continue
        lower = stripped.lower()
        if (
            stripped
            and len(stripped) <= 500
            and not any(
                label in lower
                for label in (
                    "item",
                    "description",
                    "qty",
                    "list price",
                    "discount",
                    "extended price",
                    "subtotal",
                    "grand total",
                )
            )
            and not re.fullmatch(r"[()A-Z ]+", stripped)
        ):
            pending_description = re.sub(r"\s+", " ", stripped)
        if stripped.count("$") < 3 or not re.match(r"^\d+\s", stripped):
            continue
        currency_parts = stripped.split("$")
        if len(currency_parts) != 4:
            continue
        prefix = currency_parts[0].split()
        price_discount = currency_parts[1].strip().rsplit(None, 1)
        if len(prefix) < 4 or len(price_discount) != 2:
            continue
        line_no, qty = prefix[0], prefix[-1]
        description = " ".join(prefix[1:-1]).strip()
        discount = price_discount[1].rstrip("%")
        if not line_no.isdigit() or not description:
            continue
        numeric_values = (
            qty,
            price_discount[0],
            discount,
            currency_parts[2],
            currency_parts[3],
        )
        if not all(
            re.fullmatch(r"[\d,. ()-]+", value.strip())
            for value in numeric_values
        ):
            continue
        items.append(
            ParsedQuotationItem(
                line_no=len(items) + 1,
                type="Software" if not items else "Other",
                description=description,
                qty=_number(qty),
                list_price=_number(price_discount[0]),
                discount_percent=_number(discount),
                net_unit_price=_number(currency_parts[2]),
                extended_price=_number(currency_parts[3]),
            )
        )
    return items


def _pdf_total(layout: str) -> Decimal:
    currency = r"(?:[$¥￥€]|MYR|RMB|RM)?\s*"
    for pattern in (
        rf"Grand\s+Total\s*\([^)]*\)?\s*{currency}([\d,. ]+)",
        rf"Total\s+Amount\s*:\s*{currency}([\d,. ]+)",
        rf"Grand\s+Total\s*{currency}([\d,. ]+)",
    ):
        match = re.search(pattern, layout, flags=re.IGNORECASE)
        if match:
            return _number(match.group(1))
    return Decimal("0")


def _excel_items(path: Path) -> list[ParsedQuotationItem]:
    items = []
    with path.open("rb") as stream:
        workbook = load_workbook(stream, data_only=True, read_only=True)
        try:
            for sheet in workbook.worksheets:
                price_column = None
                qty_column = None
                header_found = False
                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=min(sheet.max_row, 1000),
                    max_col=min(sheet.max_column, 50),
                    values_only=True,
                ):
                    values = [
                        "" if value is None else str(value) for value in row
                    ]
                    normalized = [value.strip().lower() for value in values]
                    if not header_found:
                        description_headers = {
                            "description",
                            "item",
                            "name/billing item",
                            "product",
                            "server",
                            "service",
                        }
                        has_description = any(
                            value in description_headers
                            for value in normalized
                        )
                        price_indexes = [
                            index
                            for index, value in enumerate(normalized)
                            if any(
                                word in value
                                for word in ("amount", "price", "cost")
                            )
                        ]
                        if not has_description or not price_indexes:
                            continue
                        price_column = price_indexes[-1]
                        qty_column = next(
                            (
                                index
                                for index, value in enumerate(normalized)
                                if value
                                in {
                                    "qty",
                                    "quantity",
                                    "subscription quantity",
                                }
                            ),
                            None,
                        )
                        header_found = True
                        continue
                    nonempty = [
                        value.strip() for value in values if value.strip()
                    ]
                    if len(nonempty) < 2:
                        continue
                    description = nonempty[0]
                    if any(
                        word in description.lower()
                        for word in ("subtotal", "grand total", "total cost")
                    ):
                        continue
                    numeric_values = []
                    for value in values:
                        if not re.search(r"\d", value):
                            continue
                        numeric_values.append(_decimal(value))
                    if not numeric_values:
                        continue
                    qty = Decimal("1")
                    if qty_column is not None and qty_column < len(row):
                        qty = _decimal(row[qty_column]) or Decimal("1")
                    price = Decimal("0")
                    if price_column is not None and price_column < len(row):
                        price = _decimal(row[price_column])
                    details = " | ".join(nonempty[:8])
                    items.append(
                        ParsedQuotationItem(
                            line_no=len(items) + 1,
                            type="Other",
                            description=details,
                            qty=qty,
                            list_price=price,
                            discount_percent=Decimal("0"),
                            net_unit_price=price,
                            extended_price=price * qty,
                        )
                    )
                    if len(items) >= 200:
                        return items
        finally:
            workbook.close()
    return items


def _excel_has_quotation_marker(path: Path) -> bool:
    markers = {
        "quotation",
        "quotation no",
        "quotation no.",
        "quote no",
        "quote no.",
    }
    with path.open("rb") as stream:
        try:
            workbook = load_workbook(stream, data_only=True, read_only=True)
        except (BadZipFile, InvalidFileException, OSError):
            return False
        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=min(sheet.max_row, 100),
                    max_col=min(sheet.max_column, 30),
                    values_only=True,
                ):
                    for value in row:
                        normalized = str(value or "").strip().lower()
                        normalized = normalized.rstrip(":")
                        if normalized in markers:
                            return True
        finally:
            workbook.close()
    return False


def _pdf_has_quotation_marker(layout: str) -> bool:
    for line in layout.splitlines():
        normalized = re.sub(r"\s+", " ", line).strip().lower()
        if normalized in {"quotation", "commercial quotation"}:
            return True
        if re.match(r"^(quotation|quote)\s+no\.?\s*:", normalized):
            return True
    return False


def _has_parsed_quotation_evidence(parsed: ParsedDocumentData) -> bool:
    quote = parsed.quotation
    confidence = parsed.field_confidence
    if quote.quote_no or confidence.get("quote_no", 0) > 0:
        return True
    return (
        bool(quote.client_company)
        and bool(quote.items)
        and confidence.get("client_company", 0) > 0
        and confidence.get("items", 0) > 0
    )


def _not_quotation_result(
    parsed: ParsedDocumentData,
) -> ParsedDocumentData:
    warnings = list(parsed.validation_warnings)
    warnings.append(
        {
            "field": "document",
            "code": "not_quotation",
            "detail": "The file does not contain quotation evidence",
        }
    )
    return ParsedDocumentData(
        quotation=parsed.quotation.model_copy(deep=True),
        document_kind="not_quotation",
        source_totals=parsed.source_totals,
        field_confidence=parsed.field_confidence,
        validation_errors=[],
        validation_warnings=warnings,
        confidence=Decimal("1.0000"),
    )


def complete_document_parse(
    asset: DocumentAsset,
    path: Path,
    parsed: ParsedDocumentData | None,
    *,
    extract_content: bool = True,
) -> ParsedDocumentData:
    """Complete partial template extraction with deterministic fallbacks."""
    if parsed is None:
        parsed = ParsedDocumentData(
            quotation=ParsedQuotation(),
            validation_errors=[
                {
                    "field": "document",
                    "code": "template_mismatch",
                    "detail": "Standard template extraction did not match",
                }
            ],
        )
    quote = parsed.quotation.model_copy(deep=True)
    fallback_used = bool(parsed.validation_errors)
    quotation_evidence = _has_parsed_quotation_evidence(parsed)
    source_date = _document_date(asset)
    layout = ""
    needs_pdf_fallback = fallback_used or not all(
        (
            quote.quote_no,
            quote.project_name,
            quote.client_company,
            quote.items,
        )
    )
    pdf_extraction_attempted = (
        asset.doc_type == "pdf" and extract_content and needs_pdf_fallback
    )
    if pdf_extraction_attempted:
        try:
            layout = _pdf_layout(path)
        except TimeoutError:
            raise
        except Exception:
            layout = ""
        quote.quote_no = quote.quote_no or re.sub(
            r"\s+",
            "",
            _pdf_label(layout, r"Quote\s+No\.?\s*:\s*([A-Za-z0-9_ .-]+)"),
        )
        quotation_evidence = quotation_evidence or bool(quote.quote_no)
        quotation_evidence = (
            quotation_evidence or _pdf_has_quotation_marker(layout)
        )
        quote.quote_date = quote.quote_date or _pdf_date(layout, "Date")
        quote.expire_date = quote.expire_date or _pdf_date(
            layout, "Quote Valid Till"
        )
        if not quotation_evidence and layout.strip():
            return _not_quotation_result(parsed)
        if not quote.items:
            quote.items = _pdf_items(layout)
        total = _pdf_total(layout)
        if not total:
            total = Decimal(parsed.source_totals.get("grand_total", "0"))
    elif asset.doc_type == "pdf":
        total = Decimal(parsed.source_totals.get("grand_total", "0"))
    elif asset.doc_type == "excel" and extract_content:
        quotation_evidence = (
            quotation_evidence or _excel_has_quotation_marker(path)
        )
        if not quotation_evidence:
            return _not_quotation_result(parsed)
        if not quote.items:
            try:
                quote.items = _excel_items(path)
            except Exception:
                quote.items = []
        total = sum(
            (item.extended_price for item in quote.items),
            Decimal("0"),
        )
    else:
        total = Decimal("0")

    if not extract_content and not quotation_evidence:
        return _not_quotation_result(parsed)

    stem = Path(asset.file_name).stem
    quote.quote_no = quote.quote_no or _safe_source_number(asset)
    quote.project_name = quote.project_name or stem[:255]
    quote.payment_terms = quote.payment_terms or "CIA"
    quote.client_company = quote.client_company or stem[:255]
    quote.contact_person = quote.contact_person or ""
    quote.email = quote.email or ""
    quote.billing_company = quote.billing_company or quote.client_company
    quote.billing_contact = quote.billing_contact or quote.contact_person
    quote.billing_email = quote.billing_email or quote.email
    quote.quote_date = quote.quote_date or source_date
    quote.expire_date = quote.expire_date or (
        quote.quote_date + timedelta(days=30)
    )
    if quote.expire_date < quote.quote_date:
        quote.expire_date = quote.quote_date + timedelta(days=30)
    quote.issuer_contact_name = (
        quote.issuer_contact_name or "OnePro Cloud Sales"
    )
    quote.issuer_contact_email = (
        quote.issuer_contact_email or "sales@oneprocloud.com"
    )
    if not quote.items:
        quote.items = [
            ParsedQuotationItem(
                line_no=1,
                type="Other",
                description=stem[:500],
                qty=Decimal("1"),
                list_price=total,
                discount_percent=Decimal("0"),
                net_unit_price=total,
                extended_price=total,
            )
        ]
    subtotal = sum(
        (item.extended_price for item in quote.items),
        Decimal("0"),
    )
    source_totals = dict(parsed.source_totals)
    source_totals.update(
        {
            "subtotal_before_vat": str(subtotal),
            "vat_amount": source_totals.get("vat_amount", "0"),
            "grand_total": str(total or subtotal),
        }
    )
    warnings = list(parsed.validation_warnings)
    if pdf_extraction_attempted and not layout.strip():
        warnings.append(
            {
                "field": "document",
                "code": "ocr_required",
                "detail": "PDF has no extractable text; OCR was queued",
            }
        )
    if fallback_used:
        warnings.append(
            {
                "field": "document",
                "code": "flexible_fallback",
                "detail": "Completed automatically using flexible extraction",
            }
        )
    return ParsedDocumentData(
        quotation=quote,
        document_kind=(
            "unknown"
            if pdf_extraction_attempted and not layout.strip()
            else "quotation"
        ),
        source_totals=source_totals,
        field_confidence=parsed.field_confidence,
        validation_errors=[],
        validation_warnings=warnings,
        confidence=max(parsed.confidence, Decimal("0.5000")),
    )
