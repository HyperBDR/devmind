from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from python_calamine import CalamineWorkbook
except ImportError:
    CalamineWorkbook = None

from quotation.services.document_parsing.schemas import (
    ParsedDocumentData,
    ParsedQuotation,
    ParsedQuotationItem,
)

PARSER_NAME = "devmind_standard_excel"
PARSER_VERSION = "2.2.0"
MONEY_TOLERANCE = Decimal("0.02")


class QuotationExcelParseError(ValueError):
    """Raised when an Excel file cannot be interpreted safely."""


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float, Decimal)):
        return str(value)
    if not isinstance(value, str) and hasattr(value, "__iter__"):
        parts = []
        for item in value:
            parts.append(str(getattr(item, "text", item)))
        return "".join(parts).strip()
    return str(value).strip()


def _normalized(value: Any) -> str:
    return (
        re.sub(r"\s+", " ", _text(value))
        .strip()
        .lower()
        .rstrip(":")
        .strip()
    )


def _decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    raw = _text(value).strip()
    negative = raw.startswith("(") and raw.endswith(")")
    raw = raw.strip("()").replace(",", "").replace("%", "")
    raw = re.sub(r"[^0-9.\-]", "", raw)
    if not raw:
        return Decimal("0")
    try:
        parsed = Decimal(raw)
    except InvalidOperation as exc:
        raise QuotationExcelParseError(
            f"Invalid numeric value: {_text(value)}"
        ) from exc
    return -parsed if negative else parsed


def _decimal_string(value: Any) -> str:
    return format(_decimal(value).normalize(), "f")


def _date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = _text(value).strip()
    if not raw:
        return None
    for candidate in (raw, raw.replace("/", "-"), raw.replace(".", "-")):
        try:
            return date.fromisoformat(candidate)
        except ValueError:
            continue
    natural = re.sub(
        r"(?<=\d)(st|nd|rd|th)\b",
        "",
        raw,
        flags=re.IGNORECASE,
    )
    for date_format in (
        "%d %B, %Y",
        "%d %b, %Y",
        "%d %B %Y",
        "%d %b %Y",
        "%d-%b-%Y",
        "%d-%b-%y",
    ):
        try:
            return datetime.strptime(natural, date_format).date()
        except ValueError:
            continue
    return None


def _rows_openpyxl(path: Path) -> list[list[Any]]:
    try:
        with path.open("rb") as stream:
            workbook = load_workbook(
                stream,
                read_only=False,
                data_only=True,
                rich_text=True,
            )
    except Exception as exc:
        raise QuotationExcelParseError(
            f"Unable to open Excel workbook: {type(exc).__name__}: {exc}"
        ) from exc
    try:
        sheet = workbook["Quotation"] if "Quotation" in workbook else None
        if sheet is None:
            sheet = workbook.worksheets[0] if workbook.worksheets else None
        if sheet is None:
            raise QuotationExcelParseError("Excel workbook has no worksheets")
        if sheet.max_row > 5000 or sheet.max_column > 200:
            raise QuotationExcelParseError("Excel worksheet is too large")
        return [
            [cell.value for cell in row]
            for row in sheet.iter_rows(
                min_row=1,
                max_row=sheet.max_row,
                max_col=min(sheet.max_column, 20),
            )
        ]
    finally:
        workbook.close()


def _rows_calamine(path: Path) -> list[list[Any]]:
    if CalamineWorkbook is None:
        raise QuotationExcelParseError("python-calamine is not installed")
    try:
        workbook = CalamineWorkbook.from_path(str(path))
        names = workbook.sheet_names
        if not names:
            raise QuotationExcelParseError("Excel workbook has no worksheets")
        sheet_name = "Quotation" if "Quotation" in names else names[0]
        rows = workbook.get_sheet_by_name(sheet_name).to_python(
            skip_empty_area=False
        )
    except QuotationExcelParseError:
        raise
    except Exception as exc:
        raise QuotationExcelParseError(
            f"Unable to stream Excel workbook: {type(exc).__name__}: {exc}"
        ) from exc
    if len(rows) > 5000:
        raise QuotationExcelParseError("Excel worksheet is too large")
    return [list(row[:20]) for row in rows]


def _rows(path: Path) -> list[list[Any]]:
    """Read using calamine when available, with openpyxl compatibility."""
    if CalamineWorkbook is not None:
        try:
            return _rows_calamine(path)
        except QuotationExcelParseError:
            pass
    return _rows_openpyxl(path)


def _starts_with_value(value: Any, label: str) -> str:
    raw = re.sub(r"\s+", " ", _text(value)).strip()
    match = re.match(
        rf"^{re.escape(label)}\s*:\s*(.*)$",
        raw,
        flags=re.IGNORECASE,
    )
    return match.group(1).strip() if match else ""


def _label_value(rows: list[list[Any]], label: str) -> str:
    target = label.lower().rstrip(":")
    for row_index, row in enumerate(rows):
        for index, value in enumerate(row):
            inline = _starts_with_value(value, label.rstrip(":"))
            if inline:
                return inline
            if _normalized(value) == target:
                for candidate in row[index + 1 :]:
                    text = _text(candidate)
                    if text:
                        return text
                for following in rows[row_index + 1 : row_index + 4]:
                    for candidate in following:
                        text = _text(candidate)
                        if text:
                            return text
    return ""


def _section_fields(
    rows: list[list[Any]], section_label: str
) -> dict[str, str]:
    start = None
    for index, row in enumerate(rows):
        if any(_normalized(value) == section_label.lower() for value in row):
            start = index + 1
            break
    if start is None:
        return {}
    result = {}
    for row in rows[start : start + 8]:
        if any(_normalized(value) in {"ship to", "bill to"} for value in row):
            break
        for value in row:
            for key, label in (
                ("company", "Company"),
                ("name", "Name"),
                ("email", "Email"),
            ):
                parsed = _starts_with_value(value, label)
                if parsed and key not in result:
                    result[key] = parsed
        for column, value in enumerate(row):
            for key, label in (
                ("company", "Company"),
                ("name", "Name"),
                ("email", "Email"),
            ):
                if key in result or _normalized(value) != label.lower():
                    continue
                for candidate in row[column + 1 :]:
                    parsed = _text(candidate)
                    if parsed:
                        result[key] = parsed
                        break
    return result


def _project_fields(rows: list[list[Any]]) -> dict[str, str]:
    headers = {
        "contact person": "issuer_contact_name",
        "email": "issuer_contact_email",
        "project": "project_name",
        "payment terms": "payment_terms",
        "currency": "currency",
    }
    for row_index, row in enumerate(rows[:-1]):
        normalized = [_normalized(value) for value in row]
        if "project" not in normalized or "currency" not in normalized:
            continue
        values = rows[row_index + 1]
        result = {}
        for column, header in enumerate(normalized):
            key = headers.get(header)
            if key and column < len(values):
                result[key] = _text(values[column])
        return result
    return {}


def _line_items(
    rows: list[list[Any]], section: str, item_type: str
) -> list[ParsedQuotationItem]:
    section_row = None
    for index, row in enumerate(rows):
        if any(_normalized(value) == section.lower() for value in row):
            section_row = index
            break
    if section_row is None:
        return []
    header_row = None
    for index in range(section_row + 1, min(section_row + 5, len(rows))):
        normalized = [_normalized(value) for value in rows[index]]
        if "description" in normalized and any(
            value == "qty" or value.startswith("qty ")
            for value in normalized
        ):
            header_row = index
            break
    if header_row is None:
        return []
    headers = [_normalized(value) for value in rows[header_row]]

    def column_for(*labels: str) -> int | None:
        for column, header in enumerate(headers):
            if any(
                header == label or header.startswith(f"{label} ")
                for label in labels
            ):
                return column
        return None

    description_column = column_for("description")
    qty_column = column_for("qty", "quantity")
    list_price_column = column_for("list price")
    discount_column = column_for("discount (%)", "discount")
    net_price_column = column_for("discounted price", "net price")
    extended_column = column_for("extended price", "amount")
    if description_column is None:
        return []

    def row_value(row: list[Any], column: int | None) -> Any:
        if column is None or column >= len(row):
            return None
        return row[column]

    def discount_value(value: Any) -> Decimal:
        parsed = _decimal(value)
        if (
            isinstance(value, (int, float, Decimal))
            and Decimal("0") <= parsed <= Decimal("1")
        ):
            return parsed * Decimal("100")
        return parsed

    items = []
    for row in rows[header_row + 1 :]:
        normalized = " ".join(_normalized(value) for value in row)
        if "subtotal" in normalized:
            break
        description = _text(row_value(row, description_column))
        if not description:
            continue
        qty_raw = row_value(row, qty_column)
        qty = _decimal(qty_raw) if _text(qty_raw) else Decimal("1")
        list_price = _decimal(row_value(row, list_price_column))
        discount = discount_value(row_value(row, discount_column))
        net_price = _decimal(row_value(row, net_price_column))
        extended = _decimal(row_value(row, extended_column))
        items.append(
            ParsedQuotationItem(
                line_no=len(items) + 1,
                type=item_type,
                description=description,
                qty=qty,
                list_price=list_price,
                discount_percent=discount,
                net_unit_price=net_price,
                extended_price=extended,
            )
        )
    return items


def _amount_by_label(rows: list[list[Any]], label: str) -> Decimal:
    target = label.lower()
    for row in rows:
        for index, value in enumerate(row):
            if target in _normalized(value):
                for candidate in reversed(row[index + 1 :]):
                    if _text(candidate):
                        return _decimal(candidate)
    return Decimal("0")


def _tax_details(rows: list[list[Any]]) -> tuple[str, Decimal]:
    for row in rows:
        for value in row:
            raw = _text(value)
            match = re.search(
                r"(.+?)\s+Amount\s*\(([0-9.]+)%\)",
                raw,
                flags=re.IGNORECASE,
            )
            if match:
                return match.group(1).strip(), Decimal(match.group(2))
    return "VAT", Decimal("0")


def _last_prefixed(rows: list[list[Any]], label: str) -> str:
    for row in reversed(rows):
        for value in reversed(row):
            parsed = _starts_with_value(value, label)
            if parsed:
                return parsed
    return ""


def _payment_term_option(payment_terms: str) -> str:
    normalized = payment_terms.upper()
    for option in ("NET 30", "NET 45", "NET 60", "CIA"):
        if option in normalized:
            return option
    return "Others" if payment_terms else "CIA"


def _issue(field: str, code: str, detail: str) -> dict:
    return {"field": field, "code": code, "detail": detail}


def _validate(
    quotation: ParsedQuotation, source_totals: dict[str, str]
) -> tuple[list[dict], list[dict]]:
    errors = []
    warnings = []
    required = (
        "quote_no",
        "project_name",
        "client_company",
        "contact_person",
        "email",
        "quote_date",
        "expire_date",
        "issuer_contact_name",
        "issuer_contact_email",
    )
    for field in required:
        if not getattr(quotation, field):
            errors.append(
                _issue(field, "required", "Required field was not parsed")
            )
    if not quotation.items:
        errors.append(
            _issue("items", "required", "No quotation line items were parsed")
        )
    if (
        quotation.quote_date
        and quotation.expire_date
        and quotation.expire_date < quotation.quote_date
    ):
        errors.append(
            _issue(
                "expire_date",
                "invalid_date_range",
                "Expiry date is before quotation date",
            )
        )
    extended_total = sum(
        (item.extended_price for item in quotation.items), Decimal("0")
    )
    subtotal = Decimal(source_totals.get("subtotal_before_vat", "0"))
    if abs(extended_total - subtotal) > MONEY_TOLERANCE:
        warnings.append(
            _issue(
                "subtotal_before_vat",
                "amount_mismatch",
                "Line item total differs from source subtotal",
            )
        )
    expected_grand = subtotal + Decimal(source_totals.get("vat_amount", "0"))
    source_grand = Decimal(source_totals.get("grand_total", "0"))
    if abs(expected_grand - source_grand) > MONEY_TOLERANCE:
        warnings.append(
            _issue(
                "grand_total",
                "amount_mismatch",
                "Source grand total differs from subtotal plus tax",
            )
        )
    return errors, warnings


def _parse_excel_rows(rows: list[list[Any]]) -> ParsedDocumentData:
    quote_no = _label_value(rows, "Quote No.")
    ship_to = _section_fields(rows, "ship to")
    bill_to = _section_fields(rows, "bill to")
    project = _project_fields(rows)
    items = _line_items(rows, "Software", "Software")
    items.extend(_line_items(rows, "Others", "Other"))
    for line_no, item in enumerate(items, start=1):
        item.line_no = line_no

    tax_label, vat_rate = _tax_details(rows)
    total_amount = _amount_by_label(rows, "total amount")
    subtotal_before_vat = _amount_by_label(rows, "subtotal before")
    grand_total = _amount_by_label(rows, "grand total")
    if not subtotal_before_vat:
        subtotal_before_vat = total_amount
    if not grand_total:
        grand_total = total_amount
    source_totals = {
        "software_subtotal": _decimal_string(
            _amount_by_label(rows, "software subscription subtotal")
        ),
        "others_subtotal": _decimal_string(
            _amount_by_label(rows, "others subtotal")
        ),
        "subtotal_before_vat": _decimal_string(subtotal_before_vat),
        "vat_amount": _decimal_string(_amount_by_label(rows, "amount (")),
        "grand_total": _decimal_string(grand_total),
    }
    payment_terms = project.get("payment_terms", "")
    issuer_company = "OnePro Cloud Limited"
    for index, row in enumerate(rows):
        if any(_normalized(value) == "quotation" for value in row):
            for previous in reversed(rows[:index]):
                values = [_text(value) for value in previous if _text(value)]
                if values:
                    issuer_company = values[0]
                    break
            break
    quotation = ParsedQuotation(
        quote_no=quote_no,
        project_name=project.get("project_name", ""),
        currency=project.get("currency", "USD").upper() or "USD",
        payment_term_option=_payment_term_option(payment_terms),
        payment_terms=payment_terms,
        quote_date=_date(_label_value(rows, "Date")),
        expire_date=_date(_label_value(rows, "Quote Valid Till")),
        tax_label=tax_label,
        vat_rate=vat_rate,
        remarks_disclaimer=_label_value(
            rows, "Additional Notes & Disclaimers"
        ),
        issuer_company_name=issuer_company,
        issuer_contact_name=project.get("issuer_contact_name", ""),
        issuer_contact_email=project.get("issuer_contact_email", ""),
        issuer_contact_title=_last_prefixed(rows, "Title"),
        client_company=ship_to.get("company", ""),
        contact_person=ship_to.get("name", ""),
        email=ship_to.get("email", ""),
        billing_company=bill_to.get("company", ""),
        billing_contact=bill_to.get("name", ""),
        billing_email=bill_to.get("email", ""),
        items=items,
    )
    errors, warnings = _validate(quotation, source_totals)
    confidence_fields = (
        "quote_no",
        "project_name",
        "client_company",
        "contact_person",
        "email",
        "quote_date",
        "expire_date",
        "issuer_contact_name",
        "issuer_contact_email",
        "payment_terms",
    )
    field_confidence = {
        field: 1.0 if getattr(quotation, field) else 0.0
        for field in confidence_fields
    }
    field_confidence["items"] = 1.0 if items else 0.0
    confidence = Decimal(
        str(sum(field_confidence.values()) / len(field_confidence))
    ).quantize(Decimal("0.0001"))
    return ParsedDocumentData(
        quotation=quotation,
        source_totals=source_totals,
        field_confidence=field_confidence,
        validation_errors=errors,
        validation_warnings=warnings,
        confidence=confidence,
    )


def parse_standard_quotation_excel(path: Path) -> ParsedDocumentData:
    """Use calamine fast path and preserve openpyxl for accurate fallback."""
    if CalamineWorkbook is not None:
        try:
            fast = _parse_excel_rows(_rows_calamine(path))
            if not fast.validation_errors and not fast.validation_warnings:
                return fast
        except QuotationExcelParseError:
            pass
    return _parse_excel_rows(_rows_openpyxl(path))
