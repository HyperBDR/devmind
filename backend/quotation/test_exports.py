import base64
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import Mock, patch

from celery.exceptions import Retry
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError, OperationalError, transaction
from django.db.models.deletion import RestrictedError
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from quotation.metrics import export_metrics_snapshot
from quotation.models import (
    EXPORT_ARCHIVE_SYNC_STAGE,
    DocumentAsset,
    DocumentParseResult,
    DocumentParseStatus,
    DocumentReplica,
    ExportJob,
    ExportJobStatus,
    Quotation,
    QuotationSourceType,
    QuotationTemplate,
    QuotationTemplateStatus,
    QuotationUploadPermission,
    QuotationVersion,
    RemoteFileCleanup,
    RemoteFileCleanupStatus,
    StorageConnection,
    StorageMount,
    SyncJob,
    SyncJobStatus,
)
from quotation.services.export_archive import (
    mark_upload_failed,
    sync_export_asset,
    update_export_upload_tracking,
)
from quotation.services.export_jobs import create_export_job
from quotation.services.export_renderer import (
    CURRENT_RENDERER_VERSION,
    PdfConversionBusyError,
    PdfConversionError,
    build_default_template_bytes,
    ensure_default_template,
    render_quotation_xlsx,
)
from quotation.services.feishu_client import FeishuAPIError
from quotation.services.quotation_service import build_quotation_snapshot
from quotation.services.storage import resolve_document_path, write_document
from quotation.services.storage_control import FeishuStorageProvider
from quotation.tasks import (
    delete_owned_remote_file_task,
    dispatch_remote_file_cleanups_task,
    render_quotation_export_task,
    sync_document_replica_task,
)


class QuotationExportFixture(TestCase):
    def setUp(self):
        self.storage = TemporaryDirectory()
        self.settings_override = override_settings(
            QUOTATION_STORAGE=self.storage.name
        )
        self.settings_override.enable()
        self.user = get_user_model().objects.create_user(
            username="export-owner@example.com",
            email="export-owner@example.com",
            password="password",
        )
        self.other_user = get_user_model().objects.create_user(
            username="other-exporter@example.com",
            email="other-exporter@example.com",
            password="password",
        )
        today = date.today()
        self.quotation = Quotation.objects.create(
            quote_no="EXPORT-001",
            project_name="Async export",
            quote_date=today,
            expire_date=today + timedelta(days=30),
            issuer_contact_name="Owner",
            issuer_contact_email=self.user.email,
            client_company="Client",
            contact_person="Contact",
            email="contact@example.com",
            created_by_email=self.user.email,
            version_current=1,
        )
        self.version = QuotationVersion.objects.create(
            quotation=self.quotation,
            version_no=1,
            status="generated",
            snapshot_json=build_quotation_snapshot(self.quotation),
            operator_email=self.user.email,
        )
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def tearDown(self):
        self.settings_override.disable()
        self.storage.cleanup()


class QuotationExportApiTests(QuotationExportFixture):

    def test_ip_allowlist_failure_explains_how_to_fix_feishu_upload(self):
        job, _created = create_export_job(
            quotation=self.quotation,
            formats=["xlsx"],
            actor=self.user,
            quotation_version_no=1,
            archive_to_feishu=True,
        )
        error = FeishuAPIError("IP is not allowed", code=99991401)

        mark_upload_failed(job.id, error)

        job.refresh_from_db()
        self.assertEqual(job.status, ExportJobStatus.UPLOAD_FAILED)
        self.assertEqual(job.error_code, "feishu_99991401")
        self.assertIn("public IPv4", job.error_message)
        self.assertIn("IP allowlist", job.error_message)

    @patch("quotation.tasks.render_quotation_export_task.apply_async")
    def test_create_export_pins_versions_and_enqueues_after_commit(
        self,
        apply_async,
    ):
        apply_async.return_value.id = "celery-task-id"

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                f"/api/v1/quotation/quotations/{self.quotation.id}/exports",
                {
                    "formats": ["xlsx", "pdf"],
                    "quotation_version": 1,
                    "archive_to_feishu": False,
                },
                format="json",
            )

        self.assertEqual(response.status_code, 202)
        job = ExportJob.objects.select_related(
            "quotation_version",
            "template",
        ).get(pk=response.data["job_id"])
        self.assertEqual(job.status, "queued")
        self.assertEqual(job.quotation_version, self.version)
        self.assertEqual(job.quotation_version_no, 1)
        self.assertEqual(job.template_version, job.template.version)
        self.assertEqual(job.formats, ["pdf", "xlsx"])
        self.assertNotEqual(job.idempotency_key, "")
        apply_async.assert_called_once_with(
            args=[job.id],
            queue="quotation_render",
        )

    @patch("quotation.tasks.render_quotation_export_task.apply_async")
    def test_repeated_request_reuses_the_same_export_job(self, apply_async):
        apply_async.return_value.id = "celery-task-id"
        payload = {
            "formats": ["xlsx"],
            "quotation_version": 1,
            "archive_to_feishu": False,
        }

        with self.captureOnCommitCallbacks(execute=True):
            first = self.client.post(
                f"/api/v1/quotation/quotations/{self.quotation.id}/exports",
                payload,
                format="json",
            )
        with self.captureOnCommitCallbacks(execute=True):
            second = self.client.post(
                f"/api/v1/quotation/quotations/{self.quotation.id}/exports",
                payload,
                format="json",
            )

        self.assertEqual(first.status_code, 202)
        self.assertEqual(second.status_code, 202)
        self.assertEqual(first.data["job_id"], second.data["job_id"])
        self.assertEqual(ExportJob.objects.count(), 1)
        apply_async.assert_called_once()

    @patch(
        "quotation.tasks.render_quotation_export_task.apply_async",
        side_effect=RuntimeError("broker unavailable"),
    )
    def test_enqueue_failure_is_visible_and_can_be_retried(self, _apply):
        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                f"/api/v1/quotation/quotations/{self.quotation.id}/exports",
                {"formats": ["xlsx"], "quotation_version": 1},
                format="json",
            )

        job = ExportJob.objects.get(pk=response.data["job_id"])
        self.assertEqual(job.status, "render_failed")
        self.assertEqual(job.error_code, "export_enqueue_failed")

    @patch("quotation.tasks.render_quotation_export_task.apply_async")
    def test_export_without_version_snapshots_current_quote(self, apply_async):
        apply_async.return_value.id = "celery-task-id"
        self.quotation.project_name = "Current project"
        self.quotation.save(update_fields=["project_name", "updated_at"])

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                f"/api/v1/quotation/quotations/{self.quotation.id}/exports",
                {"formats": ["xlsx"]},
                format="json",
            )

        job = ExportJob.objects.select_related("quotation_version").get(
            pk=response.data["job_id"]
        )
        self.assertEqual(
            job.quotation_version.snapshot_json["project_name"],
            "Current project",
        )
        self.assertEqual(job.quotation_version_no, 2)

    def test_delete_exported_quotation_cascades_jobs_and_assets(self):
        with patch("quotation.tasks.render_quotation_export_task.apply_async"):
            job, _created = create_export_job(
                quotation=self.quotation,
                formats=["xlsx"],
                actor=self.user,
                quotation_version_no=1,
            )
        render_quotation_export_task.run(job.id)
        asset = job.assets.get()
        asset_path = resolve_document_path(asset.storage_key)
        connection = StorageConnection.objects.create(
            display_name="Deletion archive",
            app_id="app-id",
            app_secret="app-secret",
        )
        mount = StorageMount.objects.create(
            connection=connection,
            root_folder_token="folder-token",
        )
        replica = DocumentReplica.objects.create(
            asset=asset,
            connection=connection,
            mount=mount,
            remote_file_token="remote-token",
            sync_status="synced",
            metadata={"remote_file_owned": True},
        )

        with self.assertRaises(RestrictedError):
            self.version.delete()
        self.assertTrue(ExportJob.objects.filter(pk=job.id).exists())
        self.assertTrue(asset_path.exists())

        with patch(
            "quotation.tasks.delete_owned_remote_file_task.apply_async"
        ) as cleanup_task:
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client.delete(
                    f"/api/v1/quotation/quotations/{self.quotation.id}"
                )

        self.assertEqual(response.status_code, 204)
        cleanup = RemoteFileCleanup.objects.get(
            remote_file_token="remote-token",
        )
        cleanup_task.assert_called_once_with(
            args=[cleanup.id],
            queue="quotation_sync",
        )
        with patch(
            "quotation.tasks.delete_owned_remote_file_task.apply_async"
        ) as duplicate_dispatch:
            dispatch_result = dispatch_remote_file_cleanups_task.run()
        self.assertEqual(dispatch_result, {"pending": 0, "dispatched": 0})
        duplicate_dispatch.assert_not_called()
        self.assertFalse(ExportJob.objects.filter(pk=job.id).exists())
        replica_exists = DocumentReplica.objects.filter(pk=replica.id).exists()
        self.assertFalse(replica_exists)
        version_exists = QuotationVersion.objects.filter(
            pk=self.version.id,
        ).exists()
        self.assertFalse(version_exists)
        self.assertFalse(asset_path.exists())

    def test_delete_quotation_rejects_active_export_job(self):
        with patch("quotation.tasks.render_quotation_export_task.apply_async"):
            job, _created = create_export_job(
                quotation=self.quotation,
                formats=["xlsx"],
                actor=self.user,
                quotation_version_no=1,
            )

        response = self.client.delete(
            f"/api/v1/quotation/quotations/{self.quotation.id}"
        )

        self.assertEqual(response.status_code, 409)
        self.assertTrue(
            Quotation.objects.filter(pk=self.quotation.pk).exists()
        )
        self.assertTrue(ExportJob.objects.filter(pk=job.pk).exists())

    def test_orm_delete_uses_central_artifact_cleanup(self):
        job, _created = create_export_job(
            quotation=self.quotation,
            formats=["xlsx"],
            actor=self.user,
            quotation_version_no=1,
        )
        render_quotation_export_task.run(job.id)
        asset_path = resolve_document_path(job.assets.get().storage_key)

        quotation_id = self.quotation.id
        with self.captureOnCommitCallbacks(execute=True):
            Quotation.objects.filter(pk=quotation_id).delete()

        self.assertFalse(asset_path.exists())
        self.assertFalse(Quotation.objects.filter(pk=quotation_id).exists())

    def test_delete_quotation_preserves_reused_remote_file(self):
        job, _created = create_export_job(
            quotation=self.quotation,
            formats=["xlsx"],
            actor=self.user,
            quotation_version_no=1,
        )
        render_quotation_export_task.run(job.id)
        asset = job.assets.get()
        connection = StorageConnection.objects.create(
            display_name="Shared archive",
            app_id="app-id",
            app_secret="app-secret",
        )
        mount = StorageMount.objects.create(
            connection=connection,
            root_folder_token="folder-token",
        )
        DocumentReplica.objects.create(
            asset=asset,
            connection=connection,
            mount=mount,
            remote_file_token="shared-token",
            sync_status="synced",
            metadata={"remote_file_owned": False},
        )

        with patch(
            "quotation.tasks.delete_owned_remote_file_task.apply_async"
        ) as cleanup_task:
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client.delete(
                    f"/api/v1/quotation/quotations/{self.quotation.id}"
                )

        self.assertEqual(response.status_code, 204)
        cleanup_task.assert_not_called()

    def test_remote_cleanup_is_queued_without_blocking_quotation_delete(self):
        job, _created = create_export_job(
            quotation=self.quotation,
            formats=["xlsx"],
            actor=self.user,
            quotation_version_no=1,
        )
        render_quotation_export_task.run(job.id)
        asset = job.assets.get()
        connection = StorageConnection.objects.create(
            display_name="Missing archive",
            app_id="app-id",
            app_secret="app-secret",
        )
        mount = StorageMount.objects.create(
            connection=connection,
            root_folder_token="folder-token",
        )
        DocumentReplica.objects.create(
            asset=asset,
            connection=connection,
            mount=mount,
            remote_file_token="missing-token",
            sync_status="synced",
            metadata={"remote_file_owned": True},
        )

        with patch(
            "quotation.tasks.delete_owned_remote_file_task.apply_async"
        ) as cleanup_task:
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client.delete(
                    f"/api/v1/quotation/quotations/{self.quotation.id}"
                )

        self.assertEqual(response.status_code, 204)
        cleanup_task.assert_called_once()
        self.assertFalse(
            Quotation.objects.filter(pk=self.quotation.id).exists()
        )

        with patch(
            "quotation.services.storage_control.FeishuStorageProvider.delete",
            side_effect=FeishuAPIError("not found", code=1061004),
        ):
            cleanup = RemoteFileCleanup.objects.get(
                remote_file_token="missing-token",
            )
            result = delete_owned_remote_file_task.run(
                cleanup.id,
            )

        self.assertEqual(result, {"status": "missing"})
        cleanup.refresh_from_db()
        self.assertEqual(cleanup.status, RemoteFileCleanupStatus.COMPLETED)

    def test_delete_quotation_preserves_file_referenced_by_other_asset(self):
        job, _created = create_export_job(
            quotation=self.quotation,
            formats=["xlsx"],
            actor=self.user,
            quotation_version_no=1,
        )
        render_quotation_export_task.run(job.id)
        connection = StorageConnection.objects.create(
            display_name="Shared reference archive",
            app_id="app-id",
            app_secret="app-secret",
        )
        mount = StorageMount.objects.create(
            connection=connection,
            root_folder_token="folder-token",
        )
        surviving_connection = StorageConnection.objects.create(
            display_name="Cross-connection shared reference",
            app_id="other-app-id",
            app_secret="other-app-secret",
        )
        surviving_mount = StorageMount.objects.create(
            connection=surviving_connection,
            root_folder_token="other-folder-token",
        )
        shared_asset = DocumentAsset.objects.create(
            doc_type="excel",
            file_name="shared.xlsx",
            mime_type="application/octet-stream",
            storage_key="documents/shared/reference",
        )
        owned_replica = DocumentReplica.objects.create(
            asset=job.assets.get(),
            connection=connection,
            mount=mount,
            remote_file_token="shared-reference-token",
            sync_status="synced",
            metadata={"remote_file_owned": True},
        )
        surviving_replica = DocumentReplica.objects.create(
            asset=shared_asset,
            connection=surviving_connection,
            mount=surviving_mount,
            remote_file_token="shared-reference-token",
            sync_status="synced",
            metadata={"remote_file_owned": False},
        )

        with patch(
            "quotation.tasks.delete_owned_remote_file_task.apply_async"
        ) as cleanup_task:
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client.delete(
                    f"/api/v1/quotation/quotations/{self.quotation.id}"
                )

        self.assertEqual(response.status_code, 204)
        cleanup = RemoteFileCleanup.objects.get(
            remote_file_token="shared-reference-token",
        )
        cleanup_task.assert_called_once_with(
            args=[cleanup.id],
            queue="quotation_sync",
        )
        self.assertFalse(
            DocumentReplica.objects.filter(pk=owned_replica.pk).exists()
        )
        with patch(
            "quotation.services.storage_control.FeishuStorageProvider.delete"
        ) as remote_delete:
            result = delete_owned_remote_file_task.run(cleanup.id)

        self.assertEqual(result, {"status": "referenced"})
        cleanup.refresh_from_db()
        surviving_replica.refresh_from_db()
        self.assertEqual(cleanup.status, RemoteFileCleanupStatus.CANCELLED)
        self.assertTrue(surviving_replica.metadata["remote_file_owned"])
        remote_delete.assert_not_called()

    def test_delete_quotation_preserves_legacy_asset_remote_reference(self):
        job, _created = create_export_job(
            quotation=self.quotation,
            formats=["xlsx"],
            actor=self.user,
            quotation_version_no=1,
        )
        render_quotation_export_task.run(job.id)
        connection = StorageConnection.objects.create(
            display_name="Legacy reference archive",
            app_id="app-id",
            app_secret="app-secret",
        )
        mount = StorageMount.objects.create(
            connection=connection,
            root_folder_token="folder-token",
        )
        DocumentReplica.objects.create(
            asset=job.assets.get(),
            connection=connection,
            mount=mount,
            remote_file_token="legacy-reference-token",
            sync_status="synced",
            metadata={"remote_file_owned": True},
        )
        DocumentAsset.objects.create(
            doc_type="excel",
            source="feishu_upload",
            feishu_file_token="legacy-reference-token",
            file_name="legacy-reference.xlsx",
            mime_type="application/octet-stream",
            storage_key="documents/legacy/reference",
        )

        with patch(
            "quotation.tasks.delete_owned_remote_file_task.apply_async"
        ) as cleanup_task:
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client.delete(
                    f"/api/v1/quotation/quotations/{self.quotation.id}"
                )

        self.assertEqual(response.status_code, 204)
        cleanup = RemoteFileCleanup.objects.get(
            remote_file_token="legacy-reference-token",
        )
        cleanup_task.assert_called_once_with(
            args=[cleanup.id],
            queue="quotation_sync",
        )
        with patch(
            "quotation.services.storage_control.FeishuStorageProvider.delete"
        ) as remote_delete:
            result = delete_owned_remote_file_task.run(cleanup.id)

        self.assertEqual(result, {"status": "referenced"})
        cleanup.refresh_from_db()
        self.assertEqual(cleanup.status, RemoteFileCleanupStatus.CANCELLED)
        remote_delete.assert_not_called()

    def test_bulk_delete_cleans_shared_owned_token_once(self):
        today = date.today()
        second_quotation = Quotation.objects.create(
            quote_no="EXPORT-002",
            project_name="Second export",
            quote_date=today,
            expire_date=today + timedelta(days=30),
            issuer_contact_name="Owner",
            issuer_contact_email=self.user.email,
            client_company="Client",
            contact_person="Contact",
            email="contact@example.com",
            created_by_email=self.user.email,
        )
        connection = StorageConnection.objects.create(
            display_name="Bulk deletion archive",
            app_id="app-id",
            app_secret="app-secret",
        )
        mount = StorageMount.objects.create(
            connection=connection,
            root_folder_token="folder-token",
        )
        for quotation, suffix in (
            (self.quotation, "one"),
            (second_quotation, "two"),
        ):
            asset = DocumentAsset.objects.create(
                quotation=quotation,
                doc_type="excel",
                file_name=f"{suffix}.xlsx",
                mime_type="application/octet-stream",
                storage_key=f"documents/bulk/{suffix}",
            )
            DocumentReplica.objects.create(
                asset=asset,
                connection=connection,
                mount=mount,
                remote_file_token="bulk-shared-token",
                sync_status="synced",
                metadata={"remote_file_owned": True},
            )

        with patch(
            "quotation.tasks.delete_owned_remote_file_task.apply_async"
        ) as cleanup_task:
            with self.captureOnCommitCallbacks(execute=True):
                Quotation.objects.filter(
                    pk__in=[self.quotation.pk, second_quotation.pk]
                ).delete()

        cleanup = RemoteFileCleanup.objects.get(
            remote_file_token="bulk-shared-token",
        )
        self.assertGreaterEqual(cleanup_task.call_count, 1)
        with patch(
            "quotation.services.storage_control.FeishuStorageProvider.delete"
        ) as remote_delete:
            first = delete_owned_remote_file_task.run(cleanup.id)
            second = delete_owned_remote_file_task.run(cleanup.id)

        self.assertEqual(first, {"status": "deleted"})
        self.assertEqual(second, {"status": "completed"})
        remote_delete.assert_called_once()

    def test_cleanup_dispatch_failure_keeps_pending_intent(self):
        asset = DocumentAsset.objects.create(
            quotation=self.quotation,
            doc_type="excel",
            file_name="broker-failure.xlsx",
            mime_type="application/octet-stream",
            storage_key="documents/broker/failure",
        )
        connection = StorageConnection.objects.create(
            display_name="Broker failure archive",
            app_id="app-id",
            app_secret="app-secret",
        )
        mount = StorageMount.objects.create(
            connection=connection,
            root_folder_token="folder-token",
        )
        DocumentReplica.objects.create(
            asset=asset,
            connection=connection,
            mount=mount,
            remote_file_token="broker-failure-token",
            sync_status="synced",
            metadata={"remote_file_owned": True},
        )

        with patch(
            "quotation.tasks.delete_owned_remote_file_task.apply_async",
            side_effect=RuntimeError("broker unavailable"),
        ):
            with self.captureOnCommitCallbacks(execute=True):
                Quotation.objects.filter(pk=self.quotation.pk).delete()

        cleanup = RemoteFileCleanup.objects.get(
            remote_file_token="broker-failure-token",
        )
        self.assertFalse(
            Quotation.objects.filter(pk=self.quotation.pk).exists()
        )
        self.assertEqual(cleanup.status, RemoteFileCleanupStatus.PENDING)
        with patch(
            "quotation.tasks.delete_owned_remote_file_task.apply_async"
        ) as retry_dispatch:
            result = dispatch_remote_file_cleanups_task.run()

        self.assertEqual(result, {"pending": 1, "dispatched": 1})
        retry_dispatch.assert_called_once_with(
            args=[cleanup.id],
            queue="quotation_sync",
        )

    def test_cleanup_dispatch_lease_advances_past_first_batch(self):
        connection = StorageConnection.objects.create(
            display_name="Cleanup lease archive",
            app_id="app-id",
            app_secret="app-secret",
        )
        RemoteFileCleanup.objects.bulk_create(
            [
                RemoteFileCleanup(
                    connection=connection,
                    remote_file_token=f"leased-token-{index}",
                    owned=True,
                )
                for index in range(201)
            ]
        )

        with patch(
            "quotation.tasks.delete_owned_remote_file_task.apply_async"
        ) as apply_async:
            first = dispatch_remote_file_cleanups_task.run()
            second = dispatch_remote_file_cleanups_task.run()

        self.assertEqual(first, {"pending": 200, "dispatched": 200})
        self.assertEqual(second, {"pending": 1, "dispatched": 1})
        self.assertEqual(apply_async.call_count, 201)

    def test_export_status_is_visible_only_to_quote_owner(self):
        response = self.client.post(
            f"/api/v1/quotation/quotations/{self.quotation.id}/exports",
            {"formats": ["xlsx"], "quotation_version": 1},
            format="json",
        )
        self.assertEqual(response.status_code, 202)

        self.client.force_authenticate(self.other_user)
        denied = self.client.get(
            f"/api/v1/quotation/exports/{response.data['job_id']}"
        )

        self.assertEqual(denied.status_code, 403)

    def test_export_rejects_unknown_quotation_version(self):
        response = self.client.post(
            f"/api/v1/quotation/quotations/{self.quotation.id}/exports",
            {"formats": ["xlsx"], "quotation_version": 99},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.data["detail"],
            "quotation version not found",
        )

    def test_staff_can_upload_and_activate_a_versioned_template(self):
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        previous = ensure_default_template(created_by=self.user)
        upload = SimpleUploadedFile(
            "quotation-v3.xlsx",
            build_default_template_bytes(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        response = self.client.post(
            "/api/v1/quotation/templates",
            {
                "name": previous.name,
                "version": 3,
                "status": "active",
                "file": upload,
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 201)
        previous.refresh_from_db()
        template = QuotationTemplate.objects.get(pk=response.data["id"])
        self.assertEqual(previous.status, "archived")
        self.assertEqual(template.status, "active")
        self.assertTrue(resolve_document_path(template.storage_key).is_file())

    def test_database_allows_only_one_active_template(self):
        ensure_default_template(created_by=self.user)
        duplicate = QuotationTemplate(
            name="Concurrent active template",
            version=1,
            storage_key="templates/concurrent.xlsx",
            content_hash="a" * 64,
            status=QuotationTemplateStatus.ACTIVE,
            created_by=self.user,
        )

        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                duplicate.save(force_insert=True)

    def test_non_staff_cannot_manage_templates(self):
        response = self.client.get("/api/v1/quotation/templates")

        self.assertEqual(response.status_code, 403)

    def test_archive_export_requires_exact_directory_upload_access(self):
        with (
            patch(
                "quotation.views.exports.feishu_common._system_drive_context",
                return_value=(Mock(), "access-token", "archive-root"),
            ),
            patch(
                "quotation.views.exports.feishu_common._managed_folder_token",
                side_effect=lambda **kwargs: kwargs["requested_token"],
            ),
        ):
            denied = self.client.post(
                f"/api/v1/quotation/quotations/{self.quotation.id}/exports",
                {
                    "formats": ["xlsx"],
                    "archive_to_feishu": True,
                    "archive_folder_token": "selected-folder",
                },
                format="json",
            )
            QuotationUploadPermission.objects.create(
                user=self.user,
                folder_token="selected-folder",
                folder_name="Selected Folder",
                granted_by=self.user,
            )
            allowed = self.client.post(
                f"/api/v1/quotation/quotations/{self.quotation.id}/exports",
                {
                    "formats": ["xlsx"],
                    "archive_to_feishu": True,
                    "archive_folder_token": "selected-folder",
                },
                format="json",
            )
            changed_token = self.client.post(
                f"/api/v1/quotation/quotations/{self.quotation.id}/exports",
                {
                    "formats": ["xlsx"],
                    "archive_to_feishu": True,
                    "archive_folder_token": "other-folder",
                },
                format="json",
            )

        self.assertEqual(denied.status_code, 403)
        self.assertEqual(
            denied.data["detail"],
            "Upload access to this directory is required.",
        )
        self.assertEqual(allowed.status_code, 202)
        self.assertEqual(changed_token.status_code, 403)


class QuotationExportTaskTests(QuotationExportFixture):
    def test_preview_excel_matches_compact_preview_layout(self):
        template = ensure_default_template()
        logo_bytes = (
            Path(__file__).resolve().parent / "assets" / "onepro-logo.png"
        ).read_bytes()
        snapshot = {
            **self.version.snapshot_json,
            "issuer_company_name": "OnePro Cloud Limited",
            "issuer_signature": (
                "data:image/png;base64,"
                + base64.b64encode(logo_bytes).decode("ascii")
            ),
            "items": [
                *[
                    {
                        "type": "Software",
                        "description": f"Software item {index}",
                        "qty": 1,
                        "list_price": 60000,
                        "discount_percent": 0,
                        "net_unit_price": 60000,
                        "extended_price": 60000,
                    }
                    for index in range(1, 5)
                ],
                *[
                    {
                        "type": "Other",
                        "description": f"Other item {index}",
                        "qty": 1,
                        "list_price": 10,
                        "discount_percent": 0,
                        "net_unit_price": 10,
                        "extended_price": 10,
                    }
                    for index in range(1, 7)
                ],
            ],
        }

        content = render_quotation_xlsx(template, snapshot)

        workbook = load_workbook(BytesIO(content))
        sheet = workbook["Quotation"]
        self.assertEqual(
            [sheet.column_dimensions[column].width for column in "ABCDEFG"],
            [12, 24, 8, 12, 10, 17, 17],
        )
        self.assertEqual(sheet["A1"].value, None)
        self.assertEqual(sheet["A2"].value, "OnePro Cloud Limited")
        self.assertEqual(sheet["A3"].value, "Quotation")
        self.assertEqual(sheet["F6"].value, "Date:")
        self.assertEqual(sheet["A7"].value, "Ship to")
        self.assertEqual(sheet["A12"].value, "Bill to:")
        self.assertEqual(sheet["A18"].value, "Contact Person")
        self.assertEqual(sheet["A21"].value, "Software")
        self.assertEqual(sheet["A22"].value, "Item")
        self.assertEqual(
            [sheet.cell(row, 2).value for row in range(23, 27)],
            [f"Software item {index}" for index in range(1, 5)],
        )
        self.assertEqual(sheet["C23"].value, 1)
        self.assertEqual(sheet["C23"].number_format, "0")
        self.assertEqual(sheet["D23"].value, 60000)
        self.assertEqual(sheet["D23"].number_format, "#,##0")
        self.assertEqual(sheet["E23"].value, 0)
        self.assertEqual(sheet["E23"].number_format, '0"%"')
        self.assertEqual(sheet["E27"].value, "Software subscription subtotal:")
        self.assertEqual(sheet["A29"].value, "Others")
        self.assertEqual(sheet["A30"].value, "Item")
        self.assertEqual(
            [sheet.cell(row, 2).value for row in range(31, 37)],
            [f"Other item {index}" for index in range(1, 7)],
        )
        self.assertEqual(len(sheet._images), 2)
        self.assertEqual(
            [type(image.anchor).__name__ for image in sheet._images],
            ["TwoCellAnchor", "TwoCellAnchor"],
        )
        workbook.close()

    def create_job(
        self,
        formats,
        *,
        archive_to_feishu=False,
        archive_folder_token="",
    ):
        resolved_folder_token = archive_folder_token
        if archive_to_feishu and not resolved_folder_token:
            resolved_folder_token = "folder-token"
        with patch("quotation.tasks.render_quotation_export_task.apply_async"):
            job, _ = create_export_job(
                quotation=self.quotation,
                formats=formats,
                actor=self.user,
                quotation_version_no=1,
                archive_to_feishu=archive_to_feishu,
                archive_folder_token=resolved_folder_token,
            )
        if archive_to_feishu:
            QuotationUploadPermission.objects.get_or_create(
                user=self.user,
                folder_token=resolved_folder_token,
                defaults={
                    "folder_name": "Archive",
                    "granted_by": self.user,
                },
            )
        return job

    def create_imported_excel_source(
        self,
        content: bytes,
        *,
        persist: bool = True,
    ) -> DocumentAsset:
        self.quotation.source_type = QuotationSourceType.DOCUMENT_IMPORT
        self.quotation.save(update_fields=["source_type", "updated_at"])
        asset = DocumentAsset.objects.create(
            quotation=self.quotation,
            doc_type="excel",
            file_name="Original quotation.xlsx",
            mime_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            storage_key="documents/original/source.xlsx",
            size_bytes=len(content),
            source="local",
            created_by_email=self.user.email,
        )
        DocumentParseResult.objects.create(
            asset=asset,
            quotation=self.quotation,
            status=DocumentParseStatus.CONFIRMED,
            parser_name="test-parser",
            parser_version="1",
            content_hash="source-content-hash",
        )
        if persist:
            write_document(content, asset.storage_key)
        return asset

    def test_imported_first_revision_uses_untouched_source_excel(self):
        source_bytes = b"PK\x03\x04-untouched-original-excel"
        self.create_imported_excel_source(source_bytes)
        job = self.create_job(["xlsx", "pdf"])

        with (
            patch(
                "quotation.services.export_pipeline.render_quotation_xlsx"
            ) as render_xlsx,
            patch(
                "quotation.services.export_pipeline.convert_xlsx_to_pdf",
                return_value=b"%PDF-from-original",
            ) as convert_pdf,
        ):
            result = render_quotation_export_task.run(job.id)

        self.assertEqual(result["status"], ExportJobStatus.COMPLETED)
        render_xlsx.assert_not_called()
        convert_pdf.assert_called_once_with(source_bytes, job_id=job.id)
        excel_asset = job.assets.get(doc_type="excel")
        self.assertEqual(
            resolve_document_path(excel_asset.storage_key).read_bytes(),
            source_bytes,
        )

    @patch(
        "quotation.services.export_pipeline.convert_xlsx_to_pdf",
        return_value=b"%PDF-preview",
    )
    @patch(
        "quotation.services.export_pipeline.render_quotation_xlsx",
        return_value=b"PK\x03\x04-preview-layout",
    )
    def test_local_pdf_converts_the_exact_preview_excel(
        self,
        render_xlsx,
        convert_pdf,
    ):
        job = self.create_job(["xlsx", "pdf"])

        result = render_quotation_export_task.run(job.id)

        self.assertEqual(result["status"], ExportJobStatus.COMPLETED)
        render_xlsx.assert_called_once_with(
            job.template,
            job.quotation_version.snapshot_json,
        )
        convert_pdf.assert_called_once_with(
            b"PK\x03\x04-preview-layout",
            job_id=job.id,
        )

    def test_imported_first_revision_fails_when_source_file_is_missing(self):
        self.create_imported_excel_source(b"missing", persist=False)
        job = self.create_job(["xlsx"])

        with patch(
            "quotation.services.export_pipeline.render_quotation_xlsx"
        ) as render_xlsx:
            result = render_quotation_export_task.run(job.id)

        job.refresh_from_db()
        self.assertEqual(result["status"], ExportJobStatus.RENDER_FAILED)
        self.assertEqual(job.error_code, "original_import_unavailable")
        self.assertEqual(job.assets.count(), 0)
        render_xlsx.assert_not_called()

    def test_imported_later_revision_still_uses_its_snapshot(self):
        self.create_imported_excel_source(b"original")
        later_version = QuotationVersion.objects.create(
            quotation=self.quotation,
            version_no=2,
            status="generated",
            snapshot_json={
                **build_quotation_snapshot(self.quotation),
                "project_name": "Edited revision",
            },
            operator_email=self.user.email,
        )
        with patch("quotation.tasks.render_quotation_export_task.apply_async"):
            job, _ = create_export_job(
                quotation=self.quotation,
                formats=["xlsx"],
                actor=self.user,
                quotation_version_no=later_version.version_no,
            )

        with patch(
            "quotation.services.export_pipeline.render_quotation_xlsx",
            return_value=b"generated-revision",
        ) as render_xlsx:
            result = render_quotation_export_task.run(job.id)

        self.assertEqual(result["status"], ExportJobStatus.COMPLETED)
        render_xlsx.assert_called_once_with(
            job.template,
            later_version.snapshot_json,
        )

    @patch(
        "quotation.services.export_pipeline.convert_xlsx_to_pdf",
        return_value=b"%PDF-rendered",
    )
    def test_render_task_persists_versioned_assets(self, _convert):
        before = export_metrics_snapshot()
        job = self.create_job(["xlsx", "pdf"])
        self.quotation.project_name = "Changed after queue"
        self.quotation.save(update_fields=["project_name", "updated_at"])

        result = render_quotation_export_task.run(job.id)

        job.refresh_from_db()
        self.assertEqual(result["status"], "completed")
        self.assertEqual(job.status, "completed")
        assets = {asset.doc_type: asset for asset in job.assets.all()}
        self.assertEqual(set(assets), {"excel", "pdf"})
        for asset in assets.values():
            self.assertEqual(asset.quotation_version, self.version)
            self.assertEqual(asset.template_version, job.template_version)
            self.assertEqual(asset.renderer_version, job.renderer_version)
            self.assertEqual(len(asset.content_hash), 64)
            self.assertTrue(resolve_document_path(asset.storage_key).is_file())
        self.assertTrue(
            resolve_document_path(assets["excel"].storage_key)
            .read_bytes()
            .startswith(b"PK\x03\x04")
        )
        self.assertTrue(
            resolve_document_path(assets["pdf"].storage_key)
            .read_bytes()
            .startswith(b"%PDF-")
        )

        render_quotation_export_task.run(job.id)
        self.assertEqual(job.assets.count(), 2)
        after = export_metrics_snapshot()
        self.assertGreater(
            after["durations"]["render"]["count"],
            before["durations"].get("render", {}).get("count", 0),
        )
        self.assertGreater(
            after["results"]["render"]["success"],
            before["results"].get("render", {}).get("success", 0),
        )

    @patch("quotation.tasks.sync_document_replica_task.apply_async")
    def test_render_queues_archive_only_after_assets_commit(self, apply_async):
        apply_async.return_value.id = "sync-task"
        job = self.create_job(["xlsx"], archive_to_feishu=True)

        with self.captureOnCommitCallbacks(execute=True):
            result = render_quotation_export_task.run(job.id)

        job.refresh_from_db()
        asset = job.assets.get()
        self.assertEqual(result["status"], "upload_queued")
        self.assertEqual(job.status, "upload_queued")
        tracker = SyncJob.objects.get(
            asset=asset,
            stage=EXPORT_ARCHIVE_SYNC_STAGE,
        )
        apply_async.assert_called_once_with(
            args=[job.id, asset.id, tracker.id],
            queue="quotation_sync",
        )

    @patch("quotation.tasks.sync_document_replica_task.apply_async")
    def test_dispatched_archive_survives_task_id_write_failure(
        self,
        apply_async,
    ):
        apply_async.return_value.id = "sync-task"
        job = self.create_job(["xlsx"], archive_to_feishu=True)

        with patch(
            "quotation.services.export_pipeline.SyncJob.objects.filter",
        ) as tracker_filter:
            tracker_filter.return_value.update.side_effect = OperationalError(
                "database unavailable"
            )
            with self.captureOnCommitCallbacks(execute=True):
                result = render_quotation_export_task.run(job.id)

        job.refresh_from_db()
        tracker = SyncJob.objects.get(
            asset=job.assets.get(),
            stage=EXPORT_ARCHIVE_SYNC_STAGE,
        )
        self.assertEqual(result["status"], ExportJobStatus.UPLOAD_QUEUED)
        self.assertEqual(job.status, ExportJobStatus.UPLOAD_QUEUED)
        self.assertEqual(tracker.celery_task_id, "")
        apply_async.assert_called_once()

    @patch(
        "quotation.services.export_pipeline.convert_xlsx_to_pdf",
        return_value=b"%PDF-rendered",
    )
    def test_partial_archive_enqueue_failure_blocks_quotation_delete(
        self,
        _convert,
    ):
        dispatched = Mock(id="dispatched-task")
        job = self.create_job(["xlsx", "pdf"], archive_to_feishu=True)

        with patch(
            "quotation.tasks.sync_document_replica_task.apply_async",
            side_effect=[dispatched, RuntimeError("broker unavailable")],
        ):
            with self.captureOnCommitCallbacks(execute=True):
                render_quotation_export_task.run(job.id)

        job.refresh_from_db()
        statuses = set(
            SyncJob.objects.filter(
                quotation=self.quotation,
                stage=EXPORT_ARCHIVE_SYNC_STAGE,
            ).values_list("status", flat=True)
        )
        response = self.client.delete(
            f"/api/v1/quotation/quotations/{self.quotation.id}"
        )
        retry_response = self.client.post(
            f"/api/v1/quotation/exports/{job.id}/retry-upload",
            {},
            format="json",
        )

        self.assertEqual(job.status, "upload_failed")
        self.assertEqual(
            statuses,
            {SyncJobStatus.QUEUED, SyncJobStatus.FAILED},
        )
        self.assertEqual(response.status_code, 409)
        self.assertEqual(retry_response.status_code, 409)
        self.assertEqual(
            retry_response.data["detail"],
            "archive uploads are still active",
        )
        self.assertTrue(
            Quotation.objects.filter(pk=self.quotation.pk).exists()
        )

    @patch(
        "quotation.tasks.sync_document_replica_task.apply_async",
        side_effect=RuntimeError("broker unavailable"),
    )
    def test_archive_enqueue_failure_keeps_rendered_assets(self, _apply):
        job = self.create_job(["xlsx"], archive_to_feishu=True)

        with self.captureOnCommitCallbacks(execute=True):
            render_quotation_export_task.run(job.id)

        job.refresh_from_db()
        self.assertEqual(job.status, "upload_failed")
        self.assertEqual(job.error_code, "archive_enqueue_failed")
        self.assertEqual(job.assets.count(), 1)
        self.assertTrue(
            resolve_document_path(job.assets.get().storage_key).is_file()
        )

    def test_redelivered_render_recovers_an_in_progress_job(self):
        job = self.create_job(["xlsx"])
        ExportJob.objects.filter(pk=job.id).update(status="rendering_excel")

        result = render_quotation_export_task.run(job.id)

        job.refresh_from_db()
        self.assertEqual(result["status"], "completed")
        self.assertEqual(job.status, "completed")
        self.assertEqual(job.assets.count(), 1)

    @patch(
        "quotation.services.export_pipeline.render_export_job",
        side_effect=PdfConversionBusyError(
            "LibreOffice conversion capacity is busy",
            code="libreoffice_busy",
        ),
    )
    def test_busy_conversion_retries_without_marking_render_failed(
        self,
        _render,
    ):
        job = self.create_job(["pdf"])

        with patch.object(
            render_quotation_export_task,
            "retry",
            side_effect=Retry(),
        ) as retry:
            with self.assertRaises(Retry):
                render_quotation_export_task.run(job.id)

        job.refresh_from_db()
        self.assertEqual(job.status, ExportJobStatus.QUEUED)
        self.assertEqual(job.error_code, "libreoffice_busy")
        retry.assert_called_once()
        self.assertEqual(retry.call_args.kwargs["args"], (job.id, 0))
        self.assertEqual(retry.call_args.kwargs["countdown"], 10)

    @patch(
        "quotation.services.export_pipeline.render_export_job",
        side_effect=PdfConversionError(
            "LibreOffice conversion failed",
            code="libreoffice_conversion_failed",
        ),
    )
    def test_conversion_failure_stops_after_one_failure_retry(
        self,
        _render,
    ):
        job = self.create_job(["pdf"])

        with patch.object(render_quotation_export_task, "retry") as retry:
            result = render_quotation_export_task.run(job.id, 1)

        job.refresh_from_db()
        self.assertEqual(result["status"], ExportJobStatus.RENDER_FAILED)
        self.assertEqual(job.status, ExportJobStatus.RENDER_FAILED)
        self.assertEqual(job.error_code, "libreoffice_conversion_failed")
        retry.assert_not_called()

    def test_queued_job_rejects_unsupported_pinned_renderer(self):
        job = self.create_job(["xlsx"])
        ExportJob.objects.filter(pk=job.id).update(
            renderer_version="openpyxl-libreoffice-v1"
        )

        result = render_quotation_export_task.run(job.id)

        job.refresh_from_db()
        self.assertEqual(result["status"], ExportJobStatus.RENDER_FAILED)
        self.assertEqual(job.status, ExportJobStatus.RENDER_FAILED)
        self.assertEqual(job.error_code, "renderer_version_unsupported")
        self.assertEqual(job.renderer_version, "openpyxl-libreoffice-v1")
        self.assertEqual(job.assets.count(), 0)

    @override_settings(QUOTATION_RENDERER_VERSION="openpyxl-libreoffice-v1")
    def test_new_job_ignores_legacy_renderer_environment_setting(self):
        job = self.create_job(["xlsx"])

        self.assertEqual(job.renderer_version, CURRENT_RENDERER_VERSION)

    @patch(
        "quotation.services.export_pipeline.convert_xlsx_to_pdf",
        side_effect=ValueError("invalid PDF"),
    )
    def test_pdf_failure_creates_no_partial_excel_asset(self, _convert):
        job = self.create_job(["xlsx", "pdf"])

        with self.assertRaises(ValueError):
            render_quotation_export_task.run(job.id)

        job.refresh_from_db()
        self.assertEqual(job.status, "render_failed")
        self.assertEqual(job.assets.count(), 0)

    def create_storage_route(self):
        connection = StorageConnection.objects.create(
            display_name="Quotation archive",
            app_id="app-id",
            app_secret="app-secret",
            status="active",
            is_default=True,
        )
        return StorageMount.objects.create(
            connection=connection,
            root_folder_token="folder-token",
            root_folder_name="Archive",
            enabled=True,
            is_default=True,
        )

    def test_provider_reuses_same_name_when_mount_policy_requires_it(self):
        mount = self.create_storage_route()
        mount.conflict_policy = "reuse"
        mount.save(update_fields=["conflict_policy", "updated_at"])
        provider = FeishuStorageProvider(mount.connection)
        existing = {
            "token": "existing-token",
            "name": "Quote.xlsx",
            "type": "file",
            "url": "https://example.test/existing",
        }

        with patch.object(provider, "access_token", return_value="token"):
            with patch.object(
                provider.client,
                "list_folder_files",
                return_value={"files": [existing], "has_more": False},
            ):
                with patch.object(
                    provider.client,
                    "download_file",
                    return_value=(b"xlsx", "application/octet-stream"),
                ):
                    with patch.object(
                        provider.client,
                        "upload_file",
                    ) as upload:
                        result = provider.upload(
                            mount,
                            file_name="Quote.xlsx",
                            content=b"xlsx",
                        )

        self.assertEqual(result["file_token"], "existing-token")
        self.assertTrue(result["reused"])
        upload.assert_not_called()

    def test_provider_does_not_reuse_same_name_with_different_content(self):
        mount = self.create_storage_route()
        mount.conflict_policy = "reuse"
        mount.save(update_fields=["conflict_policy", "updated_at"])
        provider = FeishuStorageProvider(mount.connection)
        existing = {
            "token": "existing-token",
            "name": "Quote.xlsx",
            "type": "file",
        }

        with patch.object(provider, "access_token", return_value="token"):
            with patch.object(
                provider.client,
                "list_folder_files",
                return_value={"files": [existing], "has_more": False},
            ):
                with patch.object(
                    provider.client,
                    "download_file",
                    return_value=(b"old", "application/octet-stream"),
                ):
                    with patch.object(
                        provider.client,
                        "upload_file",
                        return_value={"file_token": "new-token"},
                    ) as upload:
                        result = provider.upload(
                            mount,
                            file_name="Quote.xlsx",
                            content=b"new",
                        )

        self.assertEqual(result["file_token"], "new-token")
        self.assertEqual(
            upload.call_args.kwargs["file_name"],
            "Quote (1).xlsx",
        )

    def test_provider_skips_download_when_remote_size_differs(self):
        mount = self.create_storage_route()
        mount.conflict_policy = "reuse"
        mount.save(update_fields=["conflict_policy", "updated_at"])
        provider = FeishuStorageProvider(mount.connection)
        existing = {
            "token": "existing-token",
            "name": "Quote.xlsx",
            "type": "file",
            "size": 100000000,
        }

        with patch.object(provider, "access_token", return_value="token"):
            with patch.object(
                provider.client,
                "list_folder_files",
                return_value={"files": [existing], "has_more": False},
            ):
                with patch.object(
                    provider.client,
                    "download_file",
                ) as download:
                    with patch.object(
                        provider.client,
                        "upload_file",
                        return_value={"file_token": "new-token"},
                    ) as upload:
                        provider.upload(
                            mount,
                            file_name="Quote.xlsx",
                            content=b"new",
                        )

        download.assert_not_called()
        self.assertEqual(
            upload.call_args.kwargs["file_name"],
            "Quote (1).xlsx",
        )

    def test_provider_renames_same_name_online_document(self):
        mount = self.create_storage_route()
        mount.conflict_policy = "reuse"
        mount.save(update_fields=["conflict_policy", "updated_at"])
        provider = FeishuStorageProvider(mount.connection)
        existing = {
            "token": "online-sheet-token",
            "name": "Quote.xlsx",
            "type": "sheet",
        }

        with patch.object(provider, "access_token", return_value="token"):
            with patch.object(
                provider.client,
                "list_folder_files",
                return_value={"files": [existing], "has_more": False},
            ):
                with patch.object(
                    provider.client,
                    "download_file",
                ) as download:
                    with patch.object(
                        provider.client,
                        "upload_file",
                        return_value={"file_token": "new-token"},
                    ) as upload:
                        result = provider.upload(
                            mount,
                            file_name="Quote.xlsx",
                            content=b"xlsx",
                        )

        self.assertEqual(result["file_token"], "new-token")
        download.assert_not_called()
        self.assertEqual(
            upload.call_args.kwargs["file_name"],
            "Quote (1).xlsx",
        )

    def test_provider_renames_same_name_when_mount_policy_requires_it(self):
        mount = self.create_storage_route()
        provider = FeishuStorageProvider(mount.connection)
        files = [
            {"token": "one", "name": "Quote.xlsx", "type": "file"},
            {"token": "two", "name": "Quote (1).xlsx", "type": "file"},
        ]

        with patch.object(provider, "access_token", return_value="token"):
            with patch.object(
                provider.client,
                "list_folder_files",
                return_value={"files": files, "has_more": False},
            ):
                with patch.object(
                    provider.client,
                    "upload_file",
                    return_value={"file_token": "new-token"},
                ) as upload:
                    provider.upload(
                        mount,
                        file_name="Quote.xlsx",
                        content=b"xlsx",
                    )

        self.assertEqual(
            upload.call_args.kwargs["file_name"],
            "Quote (2).xlsx",
        )

    @patch("quotation.tasks.sync_document_replica_task.apply_async")
    def render_archived_job(self, _apply_async):
        job = self.create_job(["xlsx"], archive_to_feishu=True)
        render_quotation_export_task.run(job.id)
        return job, job.assets.get()

    @patch(
        "quotation.services.storage_control.FeishuStorageProvider.upload",
        return_value={"file_token": "remote-file", "url": "https://x"},
    )
    def test_replica_task_is_idempotent_by_content_hash(self, upload):
        self.create_storage_route()
        job, asset = self.render_archived_job()

        first = sync_document_replica_task.run(job.id, asset.id)
        second = sync_document_replica_task.run(job.id, asset.id)

        job.refresh_from_db()
        replica = DocumentReplica.objects.get(asset=asset)
        self.assertEqual(first["status"], "completed")
        self.assertEqual(second["status"], "completed")
        self.assertEqual(job.status, "completed")
        self.assertEqual(replica.sync_status, "synced")
        self.assertEqual(replica.content_hash, asset.content_hash)
        upload.assert_called_once()

    @patch(
        "quotation.services.storage_control.FeishuStorageProvider.upload",
        return_value={"file_token": "remote-file", "url": "https://x"},
    )
    def test_replica_task_uses_selected_archive_folder(self, upload):
        self.create_storage_route()
        job = self.create_job(
            ["xlsx"],
            archive_to_feishu=True,
            archive_folder_token="selected-folder",
        )
        render_quotation_export_task.run(job.id)
        asset = job.assets.get()

        sync_document_replica_task.run(job.id, asset.id)

        upload.assert_called_once()
        self.assertEqual(
            upload.call_args.kwargs["folder_token"],
            "selected-folder",
        )
        self.assertEqual(
            DocumentReplica.objects.get(asset=asset).folder_token,
            "selected-folder",
        )

    @patch(
        "quotation.services.storage_control.FeishuStorageProvider.upload",
        return_value={"file_token": "remote-file", "url": "https://x"},
    )
    def test_replica_task_rechecks_permission_before_upload(self, upload):
        self.create_storage_route()
        permission = QuotationUploadPermission.objects.create(
            user=self.user,
            folder_token="selected-folder",
            folder_name="Selected Folder",
            granted_by=self.user,
        )
        job = self.create_job(
            ["xlsx"],
            archive_to_feishu=True,
            archive_folder_token="selected-folder",
        )
        with patch("quotation.tasks.sync_document_replica_task.apply_async"):
            render_quotation_export_task.run(job.id)
        asset = job.assets.get()
        permission.is_active = False
        permission.revoked_at = timezone.now()
        permission.save(
            update_fields=["is_active", "revoked_at", "updated_at"]
        )

        result = sync_document_replica_task.run(job.id, asset.id)

        job.refresh_from_db()
        self.assertEqual(result["status"], ExportJobStatus.UPLOAD_FAILED)
        self.assertEqual(job.status, ExportJobStatus.UPLOAD_FAILED)
        upload.assert_not_called()

    @patch(
        "quotation.services.storage_control.FeishuStorageProvider.upload",
        return_value={"file_token": "remote-file", "url": "https://x"},
    )
    def test_success_tracking_database_failure_retries_task(self, _upload):
        self.create_storage_route()
        job, asset = self.render_archived_job()
        calls = 0

        def flaky_tracking_update(*args, **kwargs):
            nonlocal calls
            calls += 1
            if calls == 1:
                raise OperationalError("database unavailable")
            return update_export_upload_tracking(*args, **kwargs)

        with patch(
            "quotation.services.export_archive.update_export_upload_tracking",
            side_effect=flaky_tracking_update,
        ):
            with patch.object(
                sync_document_replica_task,
                "retry",
                side_effect=Retry(),
            ) as retry:
                with self.assertRaises(Retry):
                    sync_document_replica_task.run(job.id, asset.id)

        tracker = SyncJob.objects.get(
            asset=asset,
            stage=EXPORT_ARCHIVE_SYNC_STAGE,
        )
        self.assertEqual(tracker.status, SyncJobStatus.RETRYING)
        retry.assert_called_once()

    @patch(
        "quotation.services.storage_control.FeishuStorageProvider.upload",
        return_value={
            "file_token": "shared-new-file",
            "url": "https://x/shared",
            "reused": True,
        },
    )
    def test_replica_token_switch_does_not_inherit_old_ownership(
        self, _upload
    ):
        mount = self.create_storage_route()
        job, asset = self.render_archived_job()
        replica = DocumentReplica.objects.create(
            asset=asset,
            connection=mount.connection,
            mount=mount,
            version=1,
            remote_file_token="owned-old-file",
            content_hash="old-content",
            sync_status="synced",
            metadata={"remote_file_owned": True},
        )

        with patch(
            "quotation.tasks.delete_owned_remote_file_task.apply_async"
        ) as cleanup_task:
            with self.captureOnCommitCallbacks(execute=True):
                sync_document_replica_task.run(job.id, asset.id)

        replica.refresh_from_db()
        cleanup = RemoteFileCleanup.objects.get(
            remote_file_token="owned-old-file",
        )
        self.assertEqual(replica.remote_file_token, "shared-new-file")
        self.assertFalse(replica.metadata["remote_file_owned"])
        cleanup_task.assert_called_once_with(
            args=[cleanup.id],
            queue="quotation_sync",
        )

    @patch(
        "quotation.services.storage_control.FeishuStorageProvider.upload",
        side_effect=FeishuAPIError("permission denied", code=999),
    )
    def test_terminal_upload_failure_keeps_local_asset(self, _upload):
        self.create_storage_route()
        job, asset = self.render_archived_job()
        asset_path = resolve_document_path(asset.storage_key)

        result = sync_document_replica_task.run(job.id, asset.id)

        job.refresh_from_db()
        self.assertEqual(result["status"], "upload_failed")
        self.assertEqual(job.status, "upload_failed")
        self.assertEqual(job.error_code, "feishu_999")
        self.assertTrue(asset_path.is_file())
        self.assertEqual(job.assets.count(), 1)
        tracker = SyncJob.objects.get(
            asset=asset,
            stage=EXPORT_ARCHIVE_SYNC_STAGE,
        )
        self.assertEqual(tracker.status, SyncJobStatus.FAILED)

    @patch(
        "quotation.services.storage_control.FeishuStorageProvider.upload",
        side_effect=FeishuAPIError("permission denied", code=999),
    )
    def test_terminal_tracking_database_failure_retries_task(self, _upload):
        self.create_storage_route()
        job, asset = self.render_archived_job()
        calls = 0

        def flaky_tracking_update(*args, **kwargs):
            nonlocal calls
            calls += 1
            if calls == 1:
                raise OperationalError("database unavailable")
            return update_export_upload_tracking(*args, **kwargs)

        with patch(
            "quotation.services.export_archive.update_export_upload_tracking",
            side_effect=flaky_tracking_update,
        ):
            with patch.object(
                sync_document_replica_task,
                "retry",
                side_effect=Retry(),
            ) as retry:
                with self.assertRaises(Retry):
                    sync_document_replica_task.run(job.id, asset.id)

        job.refresh_from_db()
        tracker = SyncJob.objects.get(
            asset=asset,
            stage=EXPORT_ARCHIVE_SYNC_STAGE,
        )
        self.assertEqual(job.status, ExportJobStatus.UPLOADING)
        self.assertEqual(tracker.status, SyncJobStatus.RUNNING)
        retry.assert_called_once()

        result = sync_document_replica_task.run(job.id, asset.id)

        job.refresh_from_db()
        tracker.refresh_from_db()
        self.assertEqual(result["status"], ExportJobStatus.UPLOAD_FAILED)
        self.assertEqual(job.status, ExportJobStatus.UPLOAD_FAILED)
        self.assertEqual(tracker.status, SyncJobStatus.FAILED)

    @patch(
        "quotation.services.storage_control.FeishuStorageProvider.upload",
        return_value={"file_token": "remote-file", "url": "https://x"},
    )
    @patch(
        "quotation.services.export_pipeline.convert_xlsx_to_pdf",
        return_value=b"%PDF-rendered",
    )
    def test_successful_sibling_upload_preserves_terminal_failure(
        self,
        _convert,
        _upload,
    ):
        mount = self.create_storage_route()
        job = self.create_job(["xlsx", "pdf"], archive_to_feishu=True)
        with patch("quotation.tasks.sync_document_replica_task.apply_async"):
            render_quotation_export_task.run(job.id)
        failed_asset = job.assets.get(doc_type="excel")
        successful_asset = job.assets.get(doc_type="pdf")
        DocumentReplica.objects.create(
            asset=failed_asset,
            connection=mount.connection,
            mount=mount,
            version=1,
            sync_status="failed",
        )
        ExportJob.objects.filter(pk=job.id).update(
            status="upload_failed",
            error_code="feishu_999",
            error_message="Remote quotation archiving failed",
        )

        result = sync_export_asset(job.id, successful_asset.id)

        job.refresh_from_db()
        self.assertEqual(result["status"], "upload_failed")
        self.assertEqual(job.status, "upload_failed")
        self.assertEqual(job.error_code, "feishu_999")

    @patch("quotation.tasks.sync_document_replica_task.apply_async")
    def test_retry_upload_rechecks_directory_upload_permission(
        self,
        apply_async,
    ):
        job = self.create_job(
            ["xlsx"],
            archive_folder_token="selected-folder",
        )
        render_quotation_export_task.run(job.id)
        ExportJob.objects.filter(pk=job.id).update(
            status=ExportJobStatus.UPLOAD_FAILED,
            archive_to_feishu=True,
            error_code="feishu_999",
        )

        denied = self.client.post(
            f"/api/v1/quotation/exports/{job.id}/retry-upload",
            {},
            format="json",
        )

        QuotationUploadPermission.objects.create(
            user=self.user,
            folder_token="selected-folder",
            folder_name="Selected Folder",
            granted_by=self.user,
        )
        apply_async.return_value.id = "retry-task"
        with self.captureOnCommitCallbacks(execute=True):
            allowed = self.client.post(
                f"/api/v1/quotation/exports/{job.id}/retry-upload",
                {},
                format="json",
            )

        self.assertEqual(denied.status_code, 403)
        self.assertEqual(allowed.status_code, 202)
        apply_async.assert_called_once()

    @patch("quotation.tasks.sync_document_replica_task.apply_async")
    def test_retry_upload_requeues_existing_assets_only(self, apply_async):
        job = self.create_job(["xlsx"])
        render_quotation_export_task.run(job.id)
        asset = job.assets.get()
        ExportJob.objects.filter(pk=job.id).update(
            status="upload_failed",
            archive_to_feishu=True,
            archive_folder_token="selected-folder",
            error_code="feishu_999",
        )
        QuotationUploadPermission.objects.create(
            user=self.user,
            folder_token="selected-folder",
            folder_name="Selected Folder",
            granted_by=self.user,
        )
        other_job = self.create_job(["pdf"])
        with patch(
            "quotation.services.export_pipeline.convert_xlsx_to_pdf",
            return_value=b"%PDF-rendered",
        ):
            render_quotation_export_task.run(other_job.id)
        SyncJob.objects.create(
            job_type="upload",
            status=SyncJobStatus.RUNNING,
            quotation=self.quotation,
            asset=other_job.assets.get(),
            stage=EXPORT_ARCHIVE_SYNC_STAGE,
        )
        apply_async.return_value.id = "retry-task"

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                f"/api/v1/quotation/exports/{job.id}/retry-upload",
                {},
                format="json",
            )

        self.assertEqual(response.status_code, 202)
        self.assertEqual(response.data["status"], "upload_queued")
        tracker = SyncJob.objects.get(
            asset=asset,
            stage=EXPORT_ARCHIVE_SYNC_STAGE,
        )
        apply_async.assert_called_once_with(
            args=[job.id, asset.id, tracker.id],
            queue="quotation_sync",
        )
        self.assertEqual(job.assets.count(), 1)

    @patch("quotation.tasks.sync_document_replica_task.apply_async")
    def test_tracking_initialization_failure_leaves_upload_retryable(
        self,
        apply_async,
    ):
        job = self.create_job(["xlsx"])
        render_quotation_export_task.run(job.id)
        ExportJob.objects.filter(pk=job.id).update(
            status=ExportJobStatus.UPLOAD_FAILED,
            archive_to_feishu=True,
            archive_folder_token="selected-folder",
            error_code="feishu_999",
        )
        QuotationUploadPermission.objects.create(
            user=self.user,
            folder_token="selected-folder",
            folder_name="Selected Folder",
            granted_by=self.user,
        )

        with patch(
            "quotation.services.export_pipeline."
            "prepare_export_upload_tracking",
            side_effect=OperationalError("database unavailable"),
        ):
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client.post(
                    f"/api/v1/quotation/exports/{job.id}/retry-upload",
                    {},
                    format="json",
                )

        job.refresh_from_db()
        self.assertEqual(response.status_code, 202)
        self.assertEqual(job.status, ExportJobStatus.UPLOAD_FAILED)
        self.assertEqual(job.error_code, "archive_tracking_init_failed")
        apply_async.assert_not_called()
