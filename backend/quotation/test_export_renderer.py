import base64
import io
import subprocess
from datetime import datetime
from tempfile import TemporaryDirectory
from threading import Event, Thread
from types import SimpleNamespace
from unittest.mock import patch
from zipfile import ZIP_DEFLATED, ZipFile

from django.test import SimpleTestCase, TestCase, override_settings
from openpyxl import load_workbook
from quotation.models import QuotationTemplate
from quotation.services.export_pipeline import _attachment_pdf_bytes
from quotation.services.export_renderer import (
    DEFAULT_TEMPLATE_NAME,
    LEGACY_DEFAULT_TEMPLATE_NAME,
    PdfConversionBusyError,
    TemplateValidationError,
    _build_managed_template_bytes,
    _signature_image,
    build_default_template_bytes,
    convert_xlsx_to_pdf,
    convert_attachment_to_pdf,
    ensure_default_template,
    estimate_wrapped_lines,
    register_template_version,
    render_quotation_xlsx,
    validate_template_bytes,
)


class QuotationTemplateRendererTests(TestCase):
    def setUp(self):
        self.storage = TemporaryDirectory()
        self.settings_override = override_settings(QUOTATION_STORAGE=self.storage.name)
        self.settings_override.enable()

    def tearDown(self):
        self.settings_override.disable()
        self.storage.cleanup()

    def test_default_template_renders_snapshot_and_dynamic_rows(self):
        template = ensure_default_template()
        snapshot = {
            "quote_no": "PINNED-001",
            "quote_date": "2026-07-24",
            "expire_date": "2026-08-24",
            "client_company": "Snapshot Client",
            "contact_person": "Buyer",
            "email": "buyer@example.com",
            "billing_company": "Billing Client",
            "billing_contact": "Billing",
            "billing_email": "billing@example.com",
            "project_name": "Pinned project",
            "payment_terms": "NET 30",
            "currency": "USD",
            "remarks_disclaimer": "Immutable snapshot",
            "issuer_company_name": "OnePro Cloud Limited",
            "issuer_contact_name": "Owner",
            "issuer_contact_email": "owner@example.com",
            "tax_label": "GST",
            "vat_rate": "10.00",
            "subtotal_before_vat": "300.00",
            "vat_amount": "30.00",
            "grand_total": "330.00",
            "items": [
                {
                    "type": "Software",
                    "line_no": 1,
                    "description": "Software",
                    "qty": "1.00",
                    "list_price": "100.00",
                    "discount_percent": "0.00",
                    "net_unit_price": "100.00",
                    "extended_price": "100.00",
                },
                {
                    "type": "Other",
                    "line_no": 2,
                    "description": "Service",
                    "qty": "2.00",
                    "list_price": "100.00",
                    "discount_percent": "0.00",
                    "net_unit_price": "100.00",
                    "extended_price": "200.00",
                },
            ],
            "software_subtotal": "100.00",
            "others_subtotal": "200.00",
        }

        content = render_quotation_xlsx(template, snapshot)

        workbook = load_workbook(io.BytesIO(content), data_only=False)
        sheet = workbook["Quotation"]
        self.assertEqual(sheet["G6"].value, "PINNED-001")
        self.assertEqual(sheet["B21"].value, "Software")
        self.assertEqual(sheet["B28"].value, "Service")
        self.assertEqual(sheet["G35"].value, 300)
        self.assertEqual(sheet["G35"].number_format, "#,##0")
        self.assertEqual(sheet["G37"].value, 330)
        self.assertEqual(sheet["E36"].value, "GST Amount (10%):")
        self.assertEqual(sheet["A40"].value, "Immutable snapshot")
        notes_row = 40
        acceptance_row = next(
            cell.row
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
            and cell.value.startswith("To indicate Customer acceptance")
        )
        self.assertEqual(acceptance_row - notes_row, 3)
        self.assertTrue(
            all(
                sheet.cell(row, 1).value in (None, "")
                for row in range(notes_row + 1, acceptance_row)
            )
        )
        merged_ranges = {str(item) for item in sheet.merged_cells.ranges}
        self.assertIn("A19:G19", merged_ranges)
        self.assertIn("E35:F35", merged_ranges)
        self.assertEqual(sheet.print_area, "'Quotation'!$A$1:$G$49")

    def test_estimates_wrapped_lines_for_long_notes(self):
        notes = "This is a long paragraph that must wrap across the merged notes area."

        self.assertGreater(estimate_wrapped_lines(notes, width=24), 1)
        self.assertEqual(estimate_wrapped_lines("first\nsecond", width=24), 2)

    def test_long_notes_receive_height_for_wrapped_content(self):
        template = ensure_default_template()
        notes = (
            "This quotation is based on the information currently provided and "
            "collected during the pre-sales stage. "
            "Any changes to scope, environment, requirements, or deployment "
            "complexity may result in adjustments to pricing and delivery "
            "timelines.\n\n"
            "The Professional Services scope covers product-level installation, "
            "configuration, and deployment activities only. The customer is "
            "responsible for ensuring the target environment is fully prepared "
            "prior to implementation."
        )

        content = render_quotation_xlsx(
            template,
            {"remarks_disclaimer": notes},
        )

        workbook = load_workbook(io.BytesIO(content), data_only=False)
        sheet = workbook["Quotation"]
        notes_row = next(
            cell.row
            for row in sheet.iter_rows()
            for cell in row
            if cell.value == "Additional Notes & Disclaimers:"
        ) + 1
        self.assertGreater(sheet.row_dimensions[notes_row].height, 30)
        workbook.close()

    def test_default_template_upgrades_legacy_active_version(self):
        legacy = register_template_version(
            name=LEGACY_DEFAULT_TEMPLATE_NAME,
            version=1,
            content=_build_managed_template_bytes(version=1),
            status="active",
        )

        upgraded = ensure_default_template()

        legacy.refresh_from_db()
        self.assertEqual(legacy.status, "archived")
        self.assertEqual(upgraded.version, 2)
        self.assertEqual(upgraded.name, DEFAULT_TEMPLATE_NAME)
        content = render_quotation_xlsx(
            upgraded,
            {"tax_label": "GST", "vat_rate": "10.00"},
        )
        rendered = load_workbook(io.BytesIO(content))
        self.assertEqual(
            rendered["Quotation"]["E36"].value,
            "GST Amount (10%):",
        )
        rendered.close()

    def test_default_template_preserves_same_name_custom_version_one(self):
        workbook = load_workbook(io.BytesIO(_build_managed_template_bytes(version=1)))
        workbook["Quotation"]["A1"] = "Custom Issuer"
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()
        custom = register_template_version(
            name=LEGACY_DEFAULT_TEMPLATE_NAME,
            version=1,
            content=output.getvalue(),
            status="active",
        )

        selected = ensure_default_template()

        custom.refresh_from_db()
        self.assertEqual(selected, custom)
        self.assertEqual(custom.status, "active")
        self.assertFalse(
            QuotationTemplate.objects.filter(
                name=DEFAULT_TEMPLATE_NAME,
                version=2,
            ).exists()
        )

    def test_default_template_does_not_activate_conflicting_draft(self):
        legacy = register_template_version(
            name=LEGACY_DEFAULT_TEMPLATE_NAME,
            version=1,
            content=_build_managed_template_bytes(version=1),
            status="active",
        )
        register_template_version(
            name=DEFAULT_TEMPLATE_NAME,
            version=2,
            content=build_default_template_bytes() + b"different",
            status="draft",
        )

        with self.assertRaises(TemplateValidationError) as raised:
            ensure_default_template()

        legacy.refresh_from_db()
        self.assertEqual(
            raised.exception.code,
            "default_template_version_conflict",
        )
        self.assertEqual(legacy.status, "active")

    def test_default_template_embeds_snapshot_signature_image(self):
        template = ensure_default_template()
        png_bytes = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwC"
            "AAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
        )
        snapshot = {
            "issuer_signature": (
                "data:image/png;base64,"
                f"{base64.b64encode(png_bytes).decode('ascii')}"
            )
        }

        content = render_quotation_xlsx(template, snapshot)

        workbook = load_workbook(io.BytesIO(content), data_only=False)
        sheet = workbook["Quotation"]
        self.assertEqual(len(sheet._images), 2)

    def test_legacy_template_without_tax_ranges_remains_renderable(self):
        workbook = load_workbook(io.BytesIO(build_default_template_bytes()))
        del workbook.defined_names["tax_label"]
        del workbook.defined_names["vat_rate"]
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()
        template = register_template_version(
            name="Legacy quotation",
            version=1,
            content=output.getvalue(),
            status="active",
        )

        content = render_quotation_xlsx(
            template,
            {"tax_label": "GST", "vat_rate": "10.00"},
        )

        rendered = load_workbook(io.BytesIO(content))
        self.assertNotIn("tax_label", rendered.defined_names)
        self.assertNotIn("vat_rate", rendered.defined_names)
        rendered.close()

    def test_preview_rows_expand_without_truncating_items(self):
        template = ensure_default_template()
        items = [
            {
                "type": "Software" if index < 4 else "Other",
                "description": f"Item {index + 1}",
            }
            for index in range(10)
        ]

        content = render_quotation_xlsx(
            template,
            {"items": items},
        )

        rendered = load_workbook(io.BytesIO(content))
        sheet = rendered["Quotation"]
        self.assertEqual(sheet["B24"].value, "Item 4")
        self.assertEqual(sheet["B34"].value, "Item 10")
        self.assertEqual(sheet.row_dimensions[24].height, 24)
        self.assertEqual(sheet.row_dimensions[34].height, 24)
        rendered.close()

    def test_preview_rows_grow_for_wrapped_descriptions(self):
        template = ensure_default_template()
        content = render_quotation_xlsx(
            template,
            {
                "items": [
                    {
                        "type": "Software",
                        "description": (
                            "License Type: Hypermotion license with a valid "
                            "period of 3 months"
                        ),
                    }
                ]
            },
        )

        rendered = load_workbook(io.BytesIO(content))
        self.assertGreater(rendered["Quotation"].row_dimensions[21].height, 24)
        rendered.close()

    def test_preview_uses_the_managed_logo(self):
        template = ensure_default_template()

        content = render_quotation_xlsx(template, {"items": []})

        rendered = load_workbook(io.BytesIO(content))
        rendered_image = rendered["Quotation"]._images[0]
        self.assertEqual(len(rendered["Quotation"]._images), 1)
        self.assertEqual(rendered_image.anchor._from.row, 0)
        self.assertEqual(rendered_image.anchor._from.col, 0)
        rendered.close()

    def test_same_template_and_snapshot_produce_identical_xlsx(self):
        template = ensure_default_template()
        snapshot = {"quote_no": "DETERMINISTIC-001"}

        class FirstSaveTime(datetime):
            @classmethod
            def now(cls, tz=None):
                return cls(2026, 7, 24, 8, 0, 0, tzinfo=tz)

        class SecondSaveTime(datetime):
            @classmethod
            def now(cls, tz=None):
                return cls(2026, 7, 24, 9, 0, 0, tzinfo=tz)

        with patch(
            "openpyxl.writer.excel.datetime.datetime",
            FirstSaveTime,
        ):
            first = render_quotation_xlsx(template, snapshot)
        with patch(
            "openpyxl.writer.excel.datetime.datetime",
            SecondSaveTime,
        ):
            second = render_quotation_xlsx(template, snapshot)

        self.assertEqual(first, second)

    def test_template_validation_rejects_macros(self):
        template = ensure_default_template()
        from quotation.services.storage import resolve_document_path

        original = resolve_document_path(template.storage_key).read_bytes()
        modified = io.BytesIO(original)
        with ZipFile(modified, mode="a", compression=ZIP_DEFLATED) as archive:
            archive.writestr("xl/vbaProject.bin", b"macro")

        with self.assertRaises(TemplateValidationError) as raised:
            validate_template_bytes(modified.getvalue())

        self.assertEqual(raised.exception.code, "template_macros_forbidden")

    @override_settings(QUOTATION_MAX_TEMPLATE_EXPANDED_BYTES=1)
    def test_template_validation_rejects_excessive_expanded_size(self):
        with self.assertRaises(TemplateValidationError) as raised:
            validate_template_bytes(build_default_template_bytes())

        self.assertEqual(raised.exception.code, "template_expanded_too_large")

    @override_settings(QUOTATION_MAX_SIGNATURE_BYTES=3)
    @patch("quotation.services.export_renderer.base64.b64decode")
    def test_signature_size_is_rejected_before_base64_decode(self, decode):
        with self.assertRaises(TemplateValidationError) as raised:
            _signature_image("data:image/png;base64,AAAAAAAA")

        self.assertEqual(raised.exception.code, "signature_too_large")
        decode.assert_not_called()


class FakeLibreOfficeProcess:
    pid = 1234
    returncode = 0

    def __init__(self, command, **kwargs):
        self.command = command
        self.kwargs = kwargs

    def communicate(self, timeout):
        output_dir = self.command[self.command.index("--outdir") + 1]
        input_path = self.command[-1]
        pdf_name = input_path.rsplit("/", 1)[-1].rsplit(".", 1)[0]
        with open(f"{output_dir}/{pdf_name}.pdf", "wb") as output:
            output.write(b"%PDF-rendered")
        return b"converted", b""


class TimeoutLibreOfficeProcess(FakeLibreOfficeProcess):
    returncode = None

    def communicate(self, timeout):
        raise subprocess.TimeoutExpired(self.command, timeout)

    def wait(self, timeout=None):
        self.returncode = -15
        return self.returncode


class LibreOfficeConversionTests(SimpleTestCase):
    def setUp(self):
        self.lock_dir = TemporaryDirectory()
        self.settings_override = override_settings(
            QUOTATION_RENDER_LOCK_DIR=self.lock_dir.name,
        )
        self.settings_override.enable()

    def tearDown(self):
        self.settings_override.disable()
        self.lock_dir.cleanup()

    @patch("quotation.services.export_pipeline.convert_attachment_to_pdf")
    @patch("quotation.services.export_pipeline.resolve_document_path")
    def test_public_office_attachment_is_converted_before_merge(
        self,
        resolve_document_path,
        convert_attachment_to_pdf,
    ):
        resolve_document_path.return_value.read_bytes.return_value = (
            b"PK\x03\x04document"
        )
        convert_attachment_to_pdf.return_value = b"%PDF-converted"
        asset = SimpleNamespace(
            storage_key="documents/test",
            file_name="scope.docx",
        )

        result = _attachment_pdf_bytes(asset, "job-id")

        self.assertEqual(result, b"%PDF-converted")
        convert_attachment_to_pdf.assert_called_once_with(
            b"PK\x03\x04document",
            "scope.docx",
            job_id="job-id",
        )

    @patch("quotation.services.export_pipeline.resolve_document_path")
    def test_public_attachment_merge_rejects_unsupported_content(
        self,
        resolve_document_path,
    ):
        resolve_document_path.return_value.read_bytes.return_value = (
            b"\x89PNG\r\n\x1a\ncontent"
        )
        asset = SimpleNamespace(
            storage_key="documents/test",
            file_name="screenshot.png",
        )

        with self.assertRaisesMessage(
            TemplateValidationError,
            "Public attachment format is not supported",
        ):
            _attachment_pdf_bytes(asset, "job-id")

    @patch(
        "quotation.services.export_renderer.subprocess.Popen",
        side_effect=FakeLibreOfficeProcess,
    )
    def test_conversion_uses_isolated_profile_and_validates_pdf(self, popen):
        result = convert_xlsx_to_pdf(b"PK\x03\x04xlsx", job_id="job-1")

        self.assertEqual(result, b"%PDF-rendered")
        command = popen.call_args.args[0]
        self.assertTrue(
            any(
                argument.startswith("-env:UserInstallation=file://")
                for argument in command
            )
        )
        self.assertTrue(popen.call_args.kwargs["start_new_session"])

    @patch("quotation.services.export_renderer.os.killpg")
    @patch(
        "quotation.services.export_renderer.subprocess.Popen",
        side_effect=TimeoutLibreOfficeProcess,
    )
    def test_timeout_terminates_the_process_group(self, _popen, killpg):
        with self.assertRaises(TimeoutError):
            convert_xlsx_to_pdf(b"PK\x03\x04xlsx", job_id="job-2")

        killpg.assert_called_once()

    def test_conversion_defers_when_all_slots_are_busy(self):
        conversion_started = Event()
        release_conversion = Event()
        first_result = []
        first_error = []

        class BlockingLibreOfficeProcess(FakeLibreOfficeProcess):
            def communicate(self, timeout):
                conversion_started.set()
                release_conversion.wait(timeout=2)
                return super().communicate(timeout)

        def run_first_conversion():
            try:
                first_result.append(
                    convert_xlsx_to_pdf(
                        b"PK\x03\x04xlsx",
                        job_id="job-busy-1",
                    )
                )
            except Exception as exc:
                first_error.append(exc)

        with TemporaryDirectory() as lock_dir:
            with self.settings(
                QUOTATION_RENDER_CONCURRENCY=1,
                QUOTATION_RENDER_LOCK_DIR=lock_dir,
            ):
                with patch(
                    "quotation.services.export_renderer.subprocess.Popen",
                    side_effect=BlockingLibreOfficeProcess,
                ) as popen:
                    thread = Thread(target=run_first_conversion)
                    thread.start()
                    self.assertTrue(conversion_started.wait(timeout=1))

                    with self.assertRaises(PdfConversionBusyError):
                        convert_xlsx_to_pdf(
                            b"PK\x03\x04xlsx",
                            job_id="job-busy-2",
                        )

                    release_conversion.set()
                    thread.join(timeout=2)

        self.assertFalse(thread.is_alive())
        self.assertEqual(first_result, [b"%PDF-rendered"])
        self.assertEqual(first_error, [])
        self.assertEqual(popen.call_count, 1)
