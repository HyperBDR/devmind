import shutil
import tempfile
from io import BytesIO
from pathlib import Path
from unittest.mock import patch
from uuid import uuid4

from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError, transaction
from django.test import TestCase, override_settings
from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from rest_framework.test import APIClient

from quotation.models import (
    DocumentAsset,
    DocumentParseResult,
    Quotation,
    QuotationVersion,
    SyncJob,
    SyncJobStatus,
)
from quotation.services.document_parsing.excel_parser import (
    parse_standard_quotation_excel,
)
from quotation.services.storage import document_storage_key, write_document


def standard_quotation_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"
    sheet.cell(1, 1, "OnePro Cloud Limited")
    sheet.cell(2, 1, "Quotation")
    sheet.cell(3, 6, "Date:")
    sheet.cell(3, 7, "2026-07-20")
    sheet.cell(4, 1, "Ship to")
    sheet.cell(4, 6, "Quote No.:")
    sheet.cell(4, 7, "CloudX160726")
    bold = InlineFont(b=True)
    sheet.cell(5, 1).value = CellRichText(
        TextBlock(bold, "Company : "),
        "QA Lifecycle Company",
    )
    sheet.cell(5, 6, "Quote Valid Till:")
    sheet.cell(5, 7, "2026-08-19")
    sheet.cell(6, 1).value = CellRichText(
        TextBlock(bold, "Name : "),
        "QA Tester",
    )
    sheet.cell(7, 1).value = CellRichText(
        TextBlock(bold, "Email : "),
        "qa@example.com",
    )
    sheet.cell(9, 1, "Bill to:")
    sheet.cell(10, 1).value = CellRichText(
        TextBlock(bold, "Company : "),
        "QA Billing Company",
    )
    sheet.cell(11, 1).value = CellRichText(
        TextBlock(bold, "Name : "),
        "Billing Tester",
    )
    sheet.cell(12, 1).value = CellRichText(
        TextBlock(bold, "Email : "),
        "billing@example.com",
    )
    headers = (
        "Contact Person",
        "Email",
        "Project",
        "",
        "",
        "Payment Terms",
        "Currency",
    )
    values = (
        "Alice Chen",
        "sales@example.com",
        "Lifecycle Test",
        "",
        "",
        "NET 30",
        "USD",
    )
    for column, value in enumerate(headers, start=1):
        sheet.cell(14, column, value)
    for column, value in enumerate(values, start=1):
        sheet.cell(15, column, value)
    line_headers = (
        "Item",
        "Description",
        "Qty",
        "List Price",
        "Discount (%)",
        "Discounted Price",
        "Extended Price",
    )
    sheet.cell(17, 1, "Software")
    for column, value in enumerate(line_headers, start=1):
        sheet.cell(18, column, value)
    software_values = (1, "Cloud platform", 2, 500, "10%", 450, 900)
    for column, value in enumerate(software_values, start=1):
        sheet.cell(19, column, value)
    sheet.cell(20, 5, "Software subscription subtotal:")
    sheet.cell(20, 7, 900)
    sheet.cell(22, 1, "Others")
    for column, value in enumerate(line_headers, start=1):
        sheet.cell(23, column, value)
    other_values = (1, "Implementation service", 1, 100, "0%", 100, 100)
    for column, value in enumerate(other_values, start=1):
        sheet.cell(24, column, value)
    sheet.cell(25, 5, "Others Subtotal:")
    sheet.cell(25, 7, 100)
    sheet.cell(27, 5, "Subtotal before VAT:")
    sheet.cell(27, 7, 1000)
    sheet.cell(28, 5, "VAT Amount (8%):")
    sheet.cell(28, 7, 80)
    sheet.cell(29, 5, "Grand Total:")
    sheet.cell(29, 7, 1080)
    sheet.cell(31, 1, "Additional Notes & Disclaimers:")
    sheet.cell(32, 1, "Source quotation notes")
    sheet.cell(35, 5, "Title : Sales Manager")
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def non_quotation_inventory_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(
        ["Server", "vCPU", "Memory (GiB)", "ESSD Size", "Physical or VM"]
    )
    sheet.append(["Static Price", 4, 8, 100, "VM"])
    sheet.append(["eMIS Web 1", 16, 32, 350, "VM"])
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def split_cell_quotation_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.cell(3, 2, "OnePro Cloud Limited")
    sheet.cell(4, 2, "Quotation")
    sheet.cell(7, 11, "Date:")
    sheet.cell(7, 12, "14th May, 2026")
    sheet.cell(8, 2, "Ship to")
    sheet.cell(8, 11, "Quote No.:")
    sheet.cell(8, 12, "Motion140526_CC")
    sheet.cell(9, 2, "Company :")
    sheet.cell(9, 4, "ASL")
    sheet.cell(9, 11, "Quote Valid Till:")
    sheet.cell(9, 12, "30th Jun, 2026")
    sheet.cell(10, 2, "Name :")
    sheet.cell(10, 4, "Jacky Lee")
    sheet.cell(11, 2, "Email :")
    sheet.cell(11, 4, "jackylee@asl.com.hk")
    sheet.cell(13, 2, "Bill to:")
    sheet.cell(14, 2, "Company :")
    sheet.cell(14, 4, "ASL")
    sheet.cell(15, 2, "Name :")
    sheet.cell(15, 4, "Jacky Lee")
    sheet.cell(16, 2, "Email :")
    sheet.cell(16, 4, "jackylee@asl.com.hk")
    for column, value in {
        2: "Contact Person",
        5: "Email",
        6: "Project",
        11: "Payment Terms",
        12: "Currency",
    }.items():
        sheet.cell(19, column, value)
    for column, value in {
        2: "Carrie Chen",
        5: "carrie.chen@oneprocloud.com",
        6: "Watsons",
        11: "CIA",
        12: "HKD",
    }.items():
        sheet.cell(20, column, value)
    line_headers = {
        2: "Item",
        3: "Description",
        8: "Qty\n(ea)",
        9: "List Price\n(HKD)",
        10: "Discount (%)",
        11: "Discounted Price\n(HKD)",
        12: "Extended Price\n(HKD)",
    }
    sheet.cell(26, 2, "Software")
    for column, value in line_headers.items():
        sheet.cell(27, column, value)
    for column, value in {
        2: 1,
        3: "HyperMotion License",
        8: 23,
        9: 820,
        10: 0.2,
        11: 656,
        12: 15088,
    }.items():
        sheet.cell(28, column, value)
    sheet.cell(33, 11, "Software subscription subtotal:")
    sheet.cell(33, 12, 15088)
    sheet.cell(36, 2, "Others")
    for column, value in line_headers.items():
        sheet.cell(37, column, value)
    for column, value in {
        2: 1,
        3: "Remote Professional Service",
        9: 30000,
        10: 0,
        11: 30000,
        12: 30000,
    }.items():
        sheet.cell(38, column, value)
    sheet.cell(43, 11, "Others Subtotal:")
    sheet.cell(43, 12, 30000)
    sheet.cell(46, 11, "Total Amount:")
    sheet.cell(46, 12, 45088)
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _minimal_pdf(lines: list[str]) -> bytes:
    def escape_pdf_text(value: str) -> str:
        return (
            value.replace("\\", "\\\\")
            .replace("(", "\\(")
            .replace(")", "\\)")
        )

    content_lines = ["BT", "/F1 10 Tf", "72 760 Td"]
    for index, line in enumerate(lines):
        if index:
            content_lines.append("0 -16 Td")
        content_lines.append(f"({escape_pdf_text(line)}) Tj")
    content_lines.append("ET")
    content = "\n".join(content_lines).encode("latin-1")
    objects = [
        b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n",
        b"2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n",
        (
            b"3 0 obj\n"
            b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
            b"/Resources << /Font << /F1 4 0 R >> >> "
            b"/Contents 5 0 R >>\nendobj\n"
        ),
        (
            b"4 0 obj\n"
            b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>\n"
            b"endobj\n"
        ),
        (
            b"5 0 obj\n<< /Length "
            + str(len(content)).encode("ascii")
            + b" >>\nstream\n"
            + content
            + b"\nendstream\nendobj\n"
        ),
    ]
    output = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(output))
        output.extend(obj)
    xref_position = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.extend(
        (
            f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref_position}\n%%EOF\n"
        ).encode("ascii")
    )
    return bytes(output)


def standard_quotation_pdf(
    *,
    quote_no: str = "PDF160726",
    project_name: str = "PDF Lifecycle Test",
) -> bytes:
    return _minimal_pdf(
        [
            "OnePro Cloud Limited",
            "Quotation",
            "Date: 2026-07-20",
            f"Quote No.: {quote_no}",
            "Quote Valid Till: 2026-08-19",
            "Ship to",
            "Company : PDF Lifecycle Company",
            "Name : PDF Tester",
            "Email : pdf@example.com",
            "Bill to:",
            "Company : PDF Billing Company",
            "Name : PDF Billing Tester",
            "Email : billing-pdf@example.com",
            "Contact Person | Email | Project | Payment Terms | Currency",
            f"Alice Chen | sales@example.com | {project_name} | NET 30 | USD",
            "Software",
            "Item | Description | Qty | List Price | Discount (%) | Discounted Price | Extended Price",
            "1 | Cloud platform | 2 | 500 | 10% | 450 | 900",
            "Software subscription subtotal: 900",
            "Others",
            "Item | Description | Qty | List Price | Discount (%) | Discounted Price | Extended Price",
            "1 | Implementation service | 1 | 100 | 0% | 100 | 100",
            "Others Subtotal: 100",
            "Subtotal before VAT: 1000",
            "VAT Amount (8%): 80",
            "Grand Total: 1080",
            "Additional Notes & Disclaimers:",
            "PDF quotation notes",
            "Title : Sales Manager",
        ]
    )


class StandardQuotationExcelParserTests(TestCase):
    def test_parses_standard_export_and_validates_totals(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "quotation.xlsx"
            path.write_bytes(standard_quotation_workbook())

            parsed = parse_standard_quotation_excel(path)

        quote = parsed.quotation
        self.assertEqual(quote.quote_no, "CloudX160726")
        self.assertEqual(quote.project_name, "Lifecycle Test")
        self.assertEqual(quote.client_company, "QA Lifecycle Company")
        self.assertEqual(quote.billing_company, "QA Billing Company")
        self.assertEqual(quote.payment_term_option, "NET 30")
        self.assertEqual(quote.issuer_contact_title, "Sales Manager")
        self.assertEqual(quote.remarks_disclaimer, "Source quotation notes")
        self.assertEqual(len(quote.items), 2)
        self.assertEqual(quote.items[0].type, "Software")
        self.assertEqual(quote.items[1].type, "Other")
        self.assertEqual(parsed.source_totals["grand_total"], "1080")
        self.assertEqual(parsed.validation_errors, [])
        self.assertEqual(parsed.validation_warnings, [])
        self.assertEqual(str(parsed.confidence), "1.0000")

    def test_calamine_fast_path_retains_validated_result(self):
        from quotation.services.document_parsing import excel_parser

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "quotation.xlsx"
            path.write_bytes(standard_quotation_workbook())
            rows = excel_parser._rows_openpyxl(path)
            with patch.object(excel_parser, "CalamineWorkbook", object()):
                with patch.object(
                    excel_parser,
                    "_rows_calamine",
                    return_value=rows,
                ):
                    with patch.object(
                        excel_parser,
                        "_rows_openpyxl",
                        side_effect=AssertionError("fallback was used"),
                    ):
                        parsed = parse_standard_quotation_excel(path)

        self.assertEqual(parsed.quotation.quote_no, "CloudX160726")
        self.assertEqual(parsed.validation_errors, [])
        self.assertEqual(parsed.validation_warnings, [])

    def test_parses_split_customer_cells_and_flexible_columns(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "asl.xlsx"
            path.write_bytes(split_cell_quotation_workbook())

            parsed = parse_standard_quotation_excel(path)

        quote = parsed.quotation
        self.assertEqual(quote.quote_no, "Motion140526_CC")
        self.assertEqual(quote.client_company, "ASL")
        self.assertEqual(quote.contact_person, "Jacky Lee")
        self.assertEqual(quote.email, "jackylee@asl.com.hk")
        self.assertEqual(quote.billing_company, "ASL")
        self.assertEqual(quote.billing_contact, "Jacky Lee")
        self.assertEqual(quote.billing_email, "jackylee@asl.com.hk")
        self.assertEqual(quote.issuer_contact_name, "Carrie Chen")
        self.assertEqual(quote.project_name, "Watsons")
        self.assertEqual(quote.quote_date.isoformat(), "2026-05-14")
        self.assertEqual(quote.expire_date.isoformat(), "2026-06-30")
        self.assertEqual(len(quote.items), 2)
        self.assertEqual(str(quote.items[0].discount_percent), "20.0")
        self.assertEqual(str(quote.items[1].qty), "1")
        self.assertEqual(parsed.source_totals["grand_total"], "45088")
        self.assertEqual(parsed.validation_errors, [])
        self.assertEqual(parsed.validation_warnings, [])


class StandardQuotationPdfParserTests(TestCase):
    def test_parses_standard_text_pdf_and_validates_totals(self):
        from quotation.services.document_parsing.pdf_parser import (
            parse_standard_quotation_pdf,
        )

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "quotation.pdf"
            path.write_bytes(standard_quotation_pdf())

            parsed = parse_standard_quotation_pdf(path)

        quote = parsed.quotation
        self.assertEqual(quote.quote_no, "PDF160726")
        self.assertEqual(quote.project_name, "PDF Lifecycle Test")
        self.assertEqual(quote.client_company, "PDF Lifecycle Company")
        self.assertEqual(quote.contact_person, "PDF Tester")
        self.assertEqual(quote.payment_term_option, "NET 30")
        self.assertEqual(len(quote.items), 2)
        self.assertEqual(parsed.source_totals["grand_total"], "1080")
        self.assertEqual(parsed.validation_errors, [])

    def test_pymupdf_fast_path_retains_validated_result(self):
        from quotation.services.document_parsing import pdf_parser

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "quotation.pdf"
            path.write_bytes(standard_quotation_pdf())
            text = pdf_parser._extract_text_pypdf(path)
            with patch.object(pdf_parser, "pymupdf", object()):
                with patch.object(
                    pdf_parser,
                    "_extract_text_pymupdf",
                    return_value=text,
                ):
                    with patch.object(
                        pdf_parser,
                        "_extract_text_pypdf",
                        side_effect=AssertionError("fallback was used"),
                    ):
                        parsed = pdf_parser.parse_standard_quotation_pdf(path)

        self.assertEqual(parsed.quotation.quote_no, "PDF160726")
        self.assertEqual(parsed.validation_errors, [])
        self.assertEqual(parsed.validation_warnings, [])

    def test_parses_compact_pdf_customer_and_amount_layout(self):
        from quotation.services.document_parsing.pdf_parser import (
            parse_quotation_pdf_text,
        )

        text = "\n".join(
            [
                "OnePro Cloud Limited",
                "Quotation",
                "Date: 14th May, 2026",
                "Ship to Quote No.: Motion140526_CC",
                "Company : ASL Quote Valid Till: 30th Jun, 2026",
                "Name : Jacky Lee",
                "Email : jackylee@asl.com.hk",
                "Bill to:",
                "Company : ASL",
                "Name : Jacky Lee",
                "Email : jackylee@asl.com.hk",
                "Contact Person Email Project Payment Terms Currency",
                (
                    "Carrie Chen carrie.chen@oneprocloud.com "
                    "Watsons CIA HKD"
                ),
                "Software",
                "Item Description Qty List Price Discount Extended Price",
                "HyperMotion License",
                "1 23 $ 820.00 20% $ 656.00 $ 15,088.00",
                "Software subscription subtotal: $ 15,088.00",
                "Others",
                "Item Description Qty List Price Discount Extended Price",
                "Remote Professional Service",
                "1 $ 30,000.00 0% $ 30,000.00 $ 30,000.00",
                "Others Subtotal: $ 30,000.00",
                "Total Amount: $ 45,088.00",
            ]
        )

        parsed = parse_quotation_pdf_text(text)

        quote = parsed.quotation
        self.assertEqual(quote.quote_no, "Motion140526_CC")
        self.assertEqual(quote.client_company, "ASL")
        self.assertEqual(quote.contact_person, "Jacky Lee")
        self.assertEqual(quote.email, "jackylee@asl.com.hk")
        self.assertEqual(quote.billing_contact, "Jacky Lee")
        self.assertEqual(quote.issuer_contact_name, "Carrie Chen")
        self.assertEqual(quote.project_name, "Watsons")
        self.assertEqual(quote.currency, "HKD")
        self.assertEqual(len(quote.items), 2)
        self.assertEqual(parsed.source_totals["grand_total"], "45088.00")
        self.assertEqual(parsed.validation_errors, [])
        self.assertEqual(parsed.validation_warnings, [])

    def test_flexible_pdf_parser_bounds_malformed_long_lines(self):
        from quotation.services.document_parsing.flexible_parser import (
            _pdf_items,
        )

        malformed = "1 " + ("broken $ " * 20000)
        self.assertEqual(_pdf_items(malformed), [])

    def test_flexible_pdf_parser_uses_preceding_item_description(self):
        from quotation.services.document_parsing.flexible_parser import (
            _pdf_items,
        )

        layout = "\n".join(
            [
                "HyperBDR Backup & DR License 1 year subscription",
                "1 5 $ 420.0 15% $ 357.0 $ 1,785.0",
                "2 One Time Installation Cost 1 $ 1,150.00 "
                "0% $ 1,150.0 $ 1,150.0",
            ]
        )

        items = _pdf_items(layout)

        self.assertEqual(len(items), 2)
        self.assertEqual(
            items[0].description,
            "HyperBDR Backup & DR License 1 year subscription",
        )
        self.assertEqual(str(items[0].extended_price), "1785.0")

    def test_flexible_pdf_total_supports_business_currencies(self):
        from quotation.services.document_parsing.flexible_parser import (
            _pdf_total,
        )

        self.assertEqual(
            _pdf_total("Total Amount: MYR 3,704.00"),
            3704,
        )
        self.assertEqual(
            _pdf_total("Total Amount: ￥ 145,600.00"),
            145600,
        )


class DocumentParseEndpointTests(TestCase):
    def setUp(self):
        self.storage = tempfile.mkdtemp()
        self.override = override_settings(QUOTATION_STORAGE=self.storage)
        self.override.enable()
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="parser-owner",
            email="parser-owner@example.com",
            password="test-password",
        )
        self.other_user = user_model.objects.create_user(
            username="parser-other",
            email="parser-other@example.com",
            password="test-password",
        )
        self.api = APIClient()
        self.api.force_authenticate(self.user)
        asset_id = str(uuid4())
        storage_key = document_storage_key(asset_id)
        content = standard_quotation_workbook()
        write_document(content, storage_key)
        self.asset = DocumentAsset.objects.create(
            id=asset_id,
            doc_type="excel",
            file_name="CloudX160726.xlsx",
            mime_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            storage_key=storage_key,
            size_bytes=len(content),
            source="feishu",
            created_by_email=self.user.email,
        )

    def tearDown(self):
        self.override.disable()
        shutil.rmtree(self.storage, ignore_errors=True)

    def create_pdf_asset(self) -> DocumentAsset:
        asset_id = str(uuid4())
        storage_key = document_storage_key(asset_id)
        content = standard_quotation_pdf()
        write_document(content, storage_key)
        return DocumentAsset.objects.create(
            id=asset_id,
            doc_type="pdf",
            file_name="PDF160726.pdf",
            mime_type="application/pdf",
            storage_key=storage_key,
            size_bytes=len(content),
            source="feishu",
            created_by_email=self.user.email,
        )

    def test_parse_is_auto_created_and_idempotent(self):
        url = f"/api/v1/quotation/documents/{self.asset.id}/parse"

        first = self.api.post(url, data={}, format="json")
        second = self.api.post(url, data={}, format="json")
        fetched = self.api.get(url)

        self.assertEqual(first.status_code, 201, first.data)
        self.assertEqual(first.data["status"], "confirmed")
        self.assertFalse(first.data["reused"])
        self.assertEqual(second.status_code, 200)
        self.assertTrue(second.data["reused"])
        self.assertEqual(fetched.status_code, 200)
        self.assertEqual(fetched.data["status"], "confirmed")
        self.assertEqual(
            first.data["normalized_json"]["quote_no"],
            "CloudX160726",
        )
        self.assertEqual(DocumentParseResult.objects.count(), 1)
        self.assertEqual(SyncJob.objects.count(), 1)
        self.assertEqual(Quotation.objects.count(), 1)

    def test_flexible_parse_does_not_invent_customer_contact(self):
        from quotation.services.document_parsing.flexible_parser import (
            complete_document_parse,
        )
        from quotation.services.document_parsing.schemas import (
            ParsedDocumentData,
            ParsedQuotation,
        )

        parsed = complete_document_parse(
            self.asset,
            Path("unused.xlsx"),
            ParsedDocumentData(quotation=ParsedQuotation()),
            extract_content=False,
        )

        self.assertEqual(parsed.quotation.contact_person, "")
        self.assertEqual(parsed.quotation.email, "")
        self.assertEqual(parsed.quotation.billing_contact, "")
        self.assertEqual(parsed.quotation.billing_email, "")

    def test_inventory_excel_is_marked_not_quotation(self):
        write_document(
            non_quotation_inventory_workbook(),
            self.asset.storage_key,
        )

        response = self.api.post(
            f"/api/v1/quotation/documents/{self.asset.id}/parse",
            data={},
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["status"], "not_quotation")
        self.assertIsNone(response.data["quotation_id"])
        self.assertFalse(Quotation.objects.exists())
        list_response = self.api.get(
            "/api/v1/quotation/documents?source=feishu"
        )
        self.assertEqual(
            list_response.data[0]["parse_status"],
            "not_quotation",
        )

    def test_non_quotation_reparse_removes_derived_quote(self):
        from quotation.services.document_parsing.service import (
            parse_and_create_quotation,
        )

        old_result, _ = parse_and_create_quotation(
            self.asset,
            actor=self.user,
        )
        old_result.parser_version = "2.1.0"
        old_result.save(update_fields=["parser_version", "updated_at"])
        write_document(
            non_quotation_inventory_workbook(),
            self.asset.storage_key,
        )

        result, reused = parse_and_create_quotation(
            self.asset,
            actor=self.user,
        )

        self.assertFalse(reused)
        self.assertEqual(result.status, "not_quotation")
        self.asset.refresh_from_db()
        self.assertIsNone(self.asset.quotation_id)
        self.assertFalse(Quotation.objects.exists())
        old_result.refresh_from_db()
        self.assertEqual(old_result.status, "superseded")

    def test_cost_study_pdf_is_marked_not_quotation(self):
        asset_id = str(uuid4())
        storage_key = document_storage_key(asset_id)
        content = _minimal_pdf(
            [
                "Cost Study",
                "Cloud server sizing and internal comparison",
                "Grand Total: 231336",
            ]
        )
        write_document(content, storage_key)
        asset = DocumentAsset.objects.create(
            id=asset_id,
            doc_type="pdf",
            file_name="Cost Study.pdf",
            mime_type="application/pdf",
            storage_key=storage_key,
            size_bytes=len(content),
            source="feishu",
            created_by_email=self.user.email,
        )

        response = self.api.post(
            f"/api/v1/quotation/documents/{asset.id}/parse",
            data={},
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["status"], "not_quotation")
        self.assertIsNone(response.data["quotation_id"])
        self.assertFalse(Quotation.objects.exists())

    def test_feishu_token_is_globally_unique(self):
        self.asset.feishu_file_token = "duplicate-task-token"
        self.asset.save(update_fields=["feishu_file_token"])
        newer_id = str(uuid4())
        newer_key = document_storage_key(newer_id)
        content = standard_quotation_workbook()
        write_document(content, newer_key)
        with self.assertRaises(IntegrityError), transaction.atomic():
            DocumentAsset.objects.create(
                id=newer_id,
                doc_type="excel",
                file_name="Newer.xlsx",
                mime_type=self.asset.mime_type,
                storage_key=newer_key,
                size_bytes=len(content),
                source="feishu",
                feishu_file_token="duplicate-task-token",
                created_by_email=self.other_user.email,
            )

    def test_other_user_cannot_parse_document(self):
        self.api.force_authenticate(self.other_user)

        response = self.api.post(
            f"/api/v1/quotation/documents/{self.asset.id}/parse",
            data={},
            format="json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(DocumentParseResult.objects.exists())

    def test_confirm_reuses_auto_created_quotation(self):
        parse_response = self.api.post(
            f"/api/v1/quotation/documents/{self.asset.id}/parse",
            data={},
            format="json",
        )
        parse_result_id = parse_response.data["id"]
        confirm_url = (
            "/api/v1/quotation/document-parse-results/"
            f"{parse_result_id}/confirm"
        )

        first = self.api.post(confirm_url, data={}, format="json")
        second = self.api.post(confirm_url, data={}, format="json")

        self.assertEqual(first.status_code, 200)
        self.assertTrue(first.data["reused"])
        self.assertEqual(second.status_code, 200)
        self.assertTrue(second.data["reused"])
        self.assertEqual(Quotation.objects.count(), 1)
        quotation = Quotation.objects.get()
        self.assertEqual(quotation.quote_no, "CloudX160726")
        self.assertEqual(quotation.status, "generated")
        self.assertEqual(quotation.source_type, "document_import")
        self.assertEqual(quotation.version_current, 1)
        self.assertEqual(quotation.items.count(), 2)
        self.assertEqual(str(quotation.grand_total), "1080.00")
        self.assertEqual(quotation.version_current, 1)
        self.assertEqual(QuotationVersion.objects.count(), 1)
        self.asset.refresh_from_db()
        self.assertEqual(self.asset.quotation_id, quotation.id)
        result = DocumentParseResult.objects.get(pk=parse_result_id)
        self.assertEqual(result.status, "confirmed")
        self.assertEqual(result.quotation_id, quotation.id)

    def test_auto_create_excel_generates_visible_quote_once(self):
        from quotation.services.document_parsing.service import (
            parse_and_create_quotation,
        )

        first, first_reused = parse_and_create_quotation(
            self.asset,
            actor=self.user,
        )
        second, second_reused = parse_and_create_quotation(
            self.asset,
            actor=self.user,
        )

        self.assertFalse(first_reused)
        self.assertTrue(second_reused)
        self.assertEqual(first.id, second.id)
        self.assertEqual(first.status, "confirmed")
        self.assertEqual(Quotation.objects.count(), 1)
        quotation = Quotation.objects.get()
        self.assertEqual(quotation.quote_no, "CloudX160726")
        self.assertEqual(quotation.project_name, "Lifecycle Test")
        self.assertEqual(quotation.client_company, "QA Lifecycle Company")
        self.assertEqual(quotation.contact_person, "QA Tester")
        self.assertEqual(str(quotation.grand_total), "1080.00")
        self.assertEqual(quotation.status, "generated")
        self.asset.refresh_from_db()
        self.assertEqual(self.asset.quotation_id, quotation.id)

        list_response = self.api.get("/api/v1/quotation/quotations")
        self.assertEqual(list_response.status_code, 200)
        self.assertEqual(list_response.data["total"], 1)
        self.assertEqual(
            list_response.data["items"][0]["quote_no"], "CloudX160726"
        )
        self.assertEqual(
            len(list_response.data["items"][0]["items"]),
            2,
        )

    def test_new_parser_version_updates_existing_import_in_place(self):
        from quotation.services.document_parsing.service import (
            parse_and_create_quotation,
        )

        old_result, _ = parse_and_create_quotation(
            self.asset,
            actor=self.user,
        )
        self.asset.refresh_from_db()
        quotation = self.asset.quotation
        old_result.parser_version = "2.0.0"
        old_result.save(update_fields=["parser_version", "updated_at"])
        quotation.client_company = "Unparsed company"
        quotation.contact_person = "Not specified"
        quotation.email = "unknown@oneprocloud.invalid"
        quotation.save(
            update_fields=[
                "client_company",
                "contact_person",
                "email",
                "updated_at",
            ]
        )

        new_result, reused = parse_and_create_quotation(
            self.asset,
            actor=self.user,
        )

        self.assertTrue(reused)
        self.assertNotEqual(new_result.id, old_result.id)
        self.assertEqual(new_result.parser_version, "2.2.0")
        self.assertEqual(new_result.status, "confirmed")
        self.assertEqual(new_result.quotation_id, quotation.id)
        self.assertEqual(Quotation.objects.count(), 1)
        quotation.refresh_from_db()
        self.assertEqual(quotation.client_company, "QA Lifecycle Company")
        self.assertEqual(quotation.contact_person, "QA Tester")
        self.assertEqual(quotation.email, "qa@example.com")
        self.assertEqual(str(quotation.grand_total), "1080.00")
        self.assertEqual(quotation.version_current, 2)
        old_result.refresh_from_db()
        self.assertEqual(old_result.status, "superseded")

    def test_auto_create_pdf_generates_visible_quote(self):
        from quotation.services.document_parsing.service import (
            parse_and_create_quotation,
        )

        asset = self.create_pdf_asset()

        result, reused = parse_and_create_quotation(asset, actor=self.user)

        self.assertFalse(reused)
        self.assertEqual(result.status, "confirmed")
        quotation = Quotation.objects.get(quote_no="PDF160726")
        self.assertEqual(quotation.project_name, "PDF Lifecycle Test")
        self.assertEqual(quotation.client_company, "PDF Lifecycle Company")
        self.assertEqual(quotation.contact_person, "PDF Tester")
        self.assertEqual(str(quotation.grand_total), "1080.00")
        self.assertEqual(quotation.items.count(), 2)

    def test_amount_warning_still_creates_quotation_automatically(self):
        source = BytesIO(standard_quotation_workbook())
        workbook = load_workbook(source)
        workbook["Quotation"].cell(27, 7, 999)
        changed = BytesIO()
        workbook.save(changed)
        workbook.close()
        write_document(changed.getvalue(), self.asset.storage_key)

        response = self.api.post(
            f"/api/v1/quotation/documents/{self.asset.id}/parse",
            data={},
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["status"], "confirmed", response.data)
        self.assertTrue(response.data["validation_warnings_json"])
        self.assertEqual(Quotation.objects.count(), 1)

    def test_flexible_excel_rejects_inventory_without_quote_markers(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Server", "vCPU", "Memory", "OS"])
        sheet.append(["Web", 4, 16, "Windows Server 2022"])
        sheet.append(["Database", 8, 32, "Linux"])
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        write_document(stream.getvalue(), self.asset.storage_key)

        response = self.api.post(
            f"/api/v1/quotation/documents/{self.asset.id}/parse",
            data={},
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(
            response.data["status"],
            "not_quotation",
            response.data,
        )
        self.assertFalse(Quotation.objects.exists())

    def test_flexible_pdf_fallback_creates_quotation(self):
        asset = self.create_pdf_asset()
        content = _minimal_pdf(
            [
                "Quotation",
                "Date: 29-Apr-26",
                "Quote No.: LEGACY-001",
                "Total Amount: $1,234.50",
            ]
        )
        write_document(content, asset.storage_key)

        response = self.api.post(
            f"/api/v1/quotation/documents/{asset.id}/parse",
            data={},
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["status"], "confirmed", response.data)
        quotation = Quotation.objects.get()
        self.assertEqual(quotation.source_quote_no, "LEGACY-001")
        self.assertEqual(str(quotation.grand_total), "1234.50")

    def test_changed_document_with_same_quote_number_creates_separate_quote(self):
        from quotation.services.document_parsing.service import (
            parse_and_create_quotation,
        )

        excel_result, excel_reused = parse_and_create_quotation(
            self.asset,
            actor=self.user,
        )
        asset_id = str(uuid4())
        storage_key = document_storage_key(asset_id)
        content = standard_quotation_pdf(
            quote_no="CloudX160726",
            project_name="Lifecycle Test",
        )
        write_document(content, storage_key)
        pdf_asset = DocumentAsset.objects.create(
            id=asset_id,
            doc_type="pdf",
            file_name="CloudX160726.pdf",
            mime_type="application/pdf",
            storage_key=storage_key,
            size_bytes=len(content),
            source="feishu",
            created_by_email=self.user.email,
        )

        pdf_result, pdf_reused = parse_and_create_quotation(
            pdf_asset,
            actor=self.user,
        )

        self.assertFalse(excel_reused)
        self.assertFalse(pdf_reused)
        self.assertEqual(Quotation.objects.count(), 2)
        quotation = Quotation.objects.get(pk=excel_result.quotation_id)
        pdf_quotation = Quotation.objects.get(pk=pdf_result.quotation_id)
        self.assertEqual(pdf_result.status, "confirmed")
        self.assertNotEqual(pdf_result.quotation_id, quotation.id)
        self.assertEqual(quotation.source_quote_no, "CloudX160726")
        self.assertEqual(pdf_quotation.source_quote_no, "CloudX160726")
        self.assertNotEqual(quotation.quote_no, pdf_quotation.quote_no)
        self.assertEqual(quotation.version_current, 1)
        self.assertEqual(pdf_quotation.version_current, 1)
        self.assertEqual(pdf_quotation.client_company, "PDF Lifecycle Company")
        self.assertEqual(QuotationVersion.objects.count(), 2)
        pdf_asset.refresh_from_db()
        self.assertEqual(pdf_asset.quotation_id, pdf_quotation.id)

    def test_identical_document_copy_creates_separate_quote(self):
        from quotation.services.document_parsing.service import (
            parse_and_create_quotation,
        )

        first_result, _ = parse_and_create_quotation(
            self.asset,
            actor=self.user,
        )
        asset_id = str(uuid4())
        storage_key = document_storage_key(asset_id)
        content = standard_quotation_workbook()
        write_document(content, storage_key)
        copied_asset = DocumentAsset.objects.create(
            id=asset_id,
            doc_type="excel",
            file_name="CloudX160726-copy.xlsx",
            mime_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            storage_key=storage_key,
            size_bytes=len(content),
            source="feishu",
            created_by_email=self.user.email,
        )

        copied_result, reused = parse_and_create_quotation(
            copied_asset,
            actor=self.user,
        )

        quotation = Quotation.objects.get(quote_no="CloudX160726")
        self.assertFalse(reused)
        self.assertEqual(first_result.quotation_id, quotation.id)
        self.assertNotEqual(copied_result.quotation_id, quotation.id)
        copied_quotation = Quotation.objects.get(pk=copied_result.quotation_id)
        self.assertEqual(copied_quotation.source_quote_no, "CloudX160726")
        self.assertEqual(quotation.version_current, 1)
        self.assertEqual(copied_quotation.version_current, 1)
        self.assertEqual(QuotationVersion.objects.count(), 2)

    def test_quote_number_collision_retries_with_asset_suffix(self):
        from quotation.services.document_parsing import service

        real_build_quotation = service.build_quotation
        attempts = []

        def build_with_first_collision(*, data, items_data):
            attempts.append(data["quote_no"])
            if len(attempts) == 1:
                raise IntegrityError("simulated concurrent quote collision")
            return real_build_quotation(data=data, items_data=items_data)

        with patch.object(
            service,
            "build_quotation",
            side_effect=build_with_first_collision,
        ):
            result, reused = service.parse_and_create_quotation(
                self.asset,
                actor=self.user,
            )

        quotation = Quotation.objects.get(pk=result.quotation_id)
        self.assertFalse(reused)
        self.assertEqual(result.status, "confirmed")
        self.assertEqual(len(attempts), 2)
        self.assertEqual(attempts[0], "CloudX160726")
        self.assertIn(self.asset.id.replace("-", ""), attempts[1])
        self.assertEqual(quotation.source_quote_no, "CloudX160726")

    def test_same_quote_number_from_other_user_creates_separate_quote(self):
        self.api.post(
            f"/api/v1/quotation/documents/{self.asset.id}/parse",
            data={},
            format="json",
        )
        asset_id = str(uuid4())
        storage_key = document_storage_key(asset_id)
        content = standard_quotation_workbook()
        write_document(content, storage_key)
        other_asset = DocumentAsset.objects.create(
            id=asset_id,
            doc_type="excel",
            file_name="CloudX160726-other.xlsx",
            mime_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            storage_key=storage_key,
            size_bytes=len(content),
            source="feishu",
            created_by_email=self.other_user.email,
        )
        self.api.force_authenticate(self.other_user)

        response = self.api.post(
            f"/api/v1/quotation/documents/{other_asset.id}/parse",
            data={},
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        other_asset.refresh_from_db()
        self.assertIsNotNone(other_asset.quotation_id)
        result = DocumentParseResult.objects.get(asset=other_asset)
        self.assertEqual(result.status, "confirmed")
        self.assertEqual(Quotation.objects.count(), 2)

    def test_document_list_exposes_parse_status(self):
        self.api.post(
            f"/api/v1/quotation/documents/{self.asset.id}/parse",
            data={},
            format="json",
        )

        response = self.api.get("/api/v1/quotation/documents?source=feishu")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data[0]["parse_status"], "confirmed")
        self.assertTrue(response.data[0]["parse_result_id"])
        self.assertTrue(response.data[0]["parsed_quotation_id"])
        self.assertEqual(
            response.data[0]["parsed_quote_no"],
            "CloudX160726",
        )

    def test_imported_quotation_is_read_only(self):
        parse_response = self.api.post(
            f"/api/v1/quotation/documents/{self.asset.id}/parse",
            data={},
            format="json",
        )
        quotation_id = parse_response.data["quotation_id"]

        update_response = self.api.put(
            f"/api/v1/quotation/quotations/{quotation_id}",
            data={},
            format="json",
        )
        delete_response = self.api.delete(
            f"/api/v1/quotation/quotations/{quotation_id}"
        )
        upload_response = self.api.post(
            "/api/v1/quotation/feishu/upload",
            data={
                "quotation_id": quotation_id,
                "file": SimpleUploadedFile(
                    "CloudX160726.xlsx",
                    standard_quotation_workbook(),
                    content_type=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                ),
            },
            format="multipart",
        )

        self.assertEqual(update_response.status_code, 409)
        self.assertEqual(delete_response.status_code, 409)
        self.assertEqual(upload_response.status_code, 409)
        self.assertTrue(Quotation.objects.filter(pk=quotation_id).exists())

    def test_feishu_sync_auto_creates_excel_and_pdf_quotes(self):
        downloaded_tokens = []

        class FakeClient:
            def get_tenant_access_token(self):
                return "tenant-token"

            def list_folder_files(self, access_token, folder_token, **kwargs):
                if folder_token != "folder_token":
                    return {"files": [], "has_more": False}
                return {
                    "files": [
                        {
                            "token": "auto_xlsx",
                            "name": "CloudX160726.xlsx",
                            "type": "file",
                            "url": (
                                "https://example.feishu.cn/file/auto_xlsx"
                            ),
                        },
                        {
                            "token": "auto_pdf",
                            "name": "PDF160726.pdf",
                            "type": "file",
                            "url": "https://example.feishu.cn/file/auto_pdf",
                        },
                        {
                            "token": "temporary_xlsx",
                            "name": "~$CloudX160726.xlsx",
                            "type": "file",
                            "size": 165,
                            "url": (
                                "https://example.feishu.cn/file/"
                                "temporary_xlsx"
                            ),
                        },
                    ],
                    "has_more": False,
                }

            def download_drive_item(self, access_token, **kwargs):
                downloaded_tokens.append(kwargs["file_token"])
                if kwargs["file_token"] == "auto_xlsx":
                    return (
                        standard_quotation_workbook(),
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet",
                        "CloudX160726.xlsx",
                    )
                if kwargs["file_token"] == "temporary_xlsx":
                    return (
                        b"temporary office lock file",
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet",
                        "~$CloudX160726.xlsx",
                    )
                return (
                    standard_quotation_pdf(),
                    "application/pdf",
                    "PDF160726.pdf",
                )

        cache.delete("quotation:feishu:archive-folder-sync")
        with self.settings(
            FEISHU_APP_ID="cli_test",
            FEISHU_APP_SECRET="secret_test",
            FEISHU_WEB_BASE_URL="https://tenant.feishu.cn",
            QUOTATION_FEISHU_ARCHIVE_FOLDER_TOKEN="folder_token",
        ):
            with patch(
                "quotation.views.feishu.common._client",
                return_value=FakeClient(),
            ):
                response = self.api.post(
                    "/api/v1/quotation/feishu/sync-folder"
                )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(
            downloaded_tokens,
            ["auto_xlsx", "auto_pdf", "temporary_xlsx"],
        )
        self.assertEqual(response.data["created_count"], 3)
        self.assertEqual(response.data["parsed_count"], 2)
        self.assertEqual(response.data["created_quotation_count"], 2)
        self.assertEqual(Quotation.objects.count(), 2)
        temporary = DocumentAsset.objects.get(
            feishu_file_token="temporary_xlsx"
        )
        self.assertEqual(temporary.file_name, "~$CloudX160726.xlsx")
        self.assertFalse(temporary.parse_results.exists())
        self.assertEqual(
            set(Quotation.objects.values_list("quote_no", flat=True)),
            {"CloudX160726", "PDF160726"},
        )
        list_response = self.api.get("/api/v1/quotation/quotations")
        self.assertEqual(list_response.status_code, 200)
        self.assertEqual(list_response.data["total"], 2)
        self.assertEqual(
            {
                item["quote_no"]: {
                    "project_name": item["project_name"],
                    "client_company": item["client_company"],
                    "contact_person": item["contact_person"],
                    "grand_total": item["grand_total"],
                    "status": item["status"],
                    "source_type": item["source_type"],
                    "source_document_type": item["source_document_type"],
                    "item_count": len(item["items"]),
                }
                for item in list_response.data["items"]
            },
            {
                "CloudX160726": {
                    "project_name": "Lifecycle Test",
                    "client_company": "QA Lifecycle Company",
                    "contact_person": "QA Tester",
                    "grand_total": "1080.00",
                    "status": "generated",
                    "source_type": "document_import",
                    "source_document_type": "excel",
                    "item_count": 2,
                },
                "PDF160726": {
                    "project_name": "PDF Lifecycle Test",
                    "client_company": "PDF Lifecycle Company",
                    "contact_person": "PDF Tester",
                    "grand_total": "1080.00",
                    "status": "generated",
                    "source_type": "document_import",
                    "source_document_type": "pdf",
                    "item_count": 2,
                },
            },
        )

    def test_feishu_sync_can_enqueue_compatible_async_job(self):
        cache.delete("quotation:feishu:archive-folder-sync")
        with patch(
            "quotation.tasks.sync_feishu_folder_task.apply_async"
        ) as enqueue:
            enqueue.return_value.id = "celery-sync-task"
            response = self.api.post(
                "/api/v1/quotation/feishu/sync-folder",
                data={"async": True},
                format="json",
            )

        self.assertEqual(response.status_code, 202)
        self.assertEqual(response.data["created_count"], 0)
        self.assertEqual(response.data["sync_status"], "queued")
        job = SyncJob.objects.get(pk=response.data["sync_job_id"])
        self.assertEqual(job.status, SyncJobStatus.QUEUED)
        self.assertEqual(job.celery_task_id, "celery-sync-task")
        enqueue.assert_called_once_with(
            args=[job.id, self.user.id],
            queue="quotation_sync",
        )

        detail = self.api.get(
            f"/api/v1/quotation/feishu/sync-jobs/{job.id}"
        )
        self.assertEqual(detail.status_code, 200)
        self.assertEqual(detail.data["status"], "queued")
        cache.delete("quotation:feishu:archive-folder-sync")
