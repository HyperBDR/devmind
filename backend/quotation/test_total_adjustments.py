from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from django.contrib.auth.models import User
from django.test import SimpleTestCase, TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from quotation.services.export_renderer import render_quotation_xlsx
from quotation.services.quotation_service import calculate_totals


class QuotationTotalCalculationTests(SimpleTestCase):
    def setUp(self):
        self.items = [
            SimpleNamespace(
                type="Software",
                extended_price=Decimal("90.00"),
            )
        ]

    def test_tax_is_added_by_default(self):
        totals = calculate_totals(self.items, Decimal("8"))

        self.assertEqual(totals["vat_amount"], Decimal("7.20"))
        self.assertEqual(totals["grand_total"], Decimal("97.20"))

    def test_tax_can_be_deducted_from_total(self):
        totals = calculate_totals(
            self.items,
            Decimal("8"),
            "subtract",
        )

        self.assertEqual(totals["vat_amount"], Decimal("7.20"))
        self.assertEqual(totals["grand_total"], Decimal("82.80"))

    def test_export_renders_deducted_tax_and_additional_total(self):
        content = render_quotation_xlsx(
            None,
            {
                "currency": "USD",
                "tax_label": "VAT",
                "tax_calculation": "subtract",
                "vat_rate": "8.00",
                "subtotal_before_vat": "90.00",
                "vat_amount": "7.20",
                "grand_total": "82.80",
                "additional_grand_total_label": "Total Payable",
                "additional_grand_total_currency": "MYR",
                "additional_grand_total_amount": "380.50",
                "items": [],
            },
        )

        workbook = load_workbook(BytesIO(content), data_only=False)
        sheet = workbook["Quotation"]
        labels = {
            sheet.cell(row, 5).value: sheet.cell(row, 7).value
            for row in range(1, sheet.max_row + 1)
        }
        self.assertEqual(labels["VAT Amount (8%):"], -7.2)
        self.assertEqual(labels["Total Payable (MYR):"], 380.5)
        self.assertEqual(labels["Grand Total:"], 82.8)
        workbook.close()


class QuotationTotalPersistenceTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(
            username="total-options-user",
            email="total-options@example.com",
            password="password",
        )
        self.api = APIClient()
        self.api.force_authenticate(user=user)

    def test_create_persists_total_options(self):
        response = self.api.post(
            "/api/v1/quotation/quotations",
            {
                "numbering_mode": "auto",
                "product_line": "BDR",
                "project_name": "Tax direction project",
                "currency": "USD",
                "payment_term_option": "CIA",
                "payment_terms": "CIA",
                "quote_date": "2026-09-03",
                "expire_date": "2026-10-03",
                "tax_label": "VAT",
                "tax_calculation": "subtract",
                "vat_rate": "8.00",
                "additional_grand_total_label": "Total Payable",
                "additional_grand_total_currency": "MYR",
                "additional_grand_total_amount": "380.50",
                "issuer_contact_name": "Carrol Yu",
                "issuer_contact_email": "carrol@example.com",
                "client_company": "Example Client",
                "contact_person": "Buyer",
                "email": "buyer@example.com",
                "items": [
                    {
                        "line_no": 1,
                        "type": "Software",
                        "description": "License",
                        "qty": "1",
                        "list_price": "90.00",
                        "discount_percent": "0",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["tax_calculation"], "subtract")
        self.assertEqual(response.data["vat_amount"], "7.20")
        self.assertEqual(response.data["grand_total"], "82.80")
        self.assertEqual(
            response.data["additional_grand_total_label"],
            "Total Payable",
        )
        self.assertEqual(
            response.data["additional_grand_total_currency"],
            "MYR",
        )
        self.assertEqual(
            response.data["additional_grand_total_amount"],
            "380.50",
        )
